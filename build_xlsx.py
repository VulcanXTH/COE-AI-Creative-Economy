"""สร้างไฟล์ Excel: COE AI Creative Economy
- Sheet 1: รายชื่อคณะทำงาน 17 คน
- Sheet 2: เป้าหมาย / Output / Outcomes 1-3-5 ปี
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THAI_FONT = "TH Sarabun New"
FALLBACK_FONT = "Arial"  # ใช้ Arial เป็น fallback บนเครื่องที่ไม่มี TH Sarabun

PRIMARY_FILL = PatternFill("solid", start_color="1F4E78")   # navy
ACCENT_FILL  = PatternFill("solid", start_color="2E75B6")   # blue
HEADER_FILL  = PatternFill("solid", start_color="BDD7EE")   # light blue
Y1_FILL      = PatternFill("solid", start_color="E2EFDA")   # light green
Y3_FILL      = PatternFill("solid", start_color="FFF2CC")   # light yellow
Y5_FILL      = PatternFill("solid", start_color="FCE4D6")   # light orange
ZEBRA_FILL   = PatternFill("solid", start_color="F2F2F2")   # light gray

thin = Side(border_style="thin", color="9BA3AF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_title(cell):
    cell.font = Font(name=FALLBACK_FONT, size=16, bold=True, color="FFFFFF")
    cell.fill = PRIMARY_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_subtitle(cell):
    cell.font = Font(name=FALLBACK_FONT, size=11, italic=True, color="FFFFFF")
    cell.fill = ACCENT_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_header(cell):
    cell.font = Font(name=FALLBACK_FONT, size=11, bold=True, color="1F2937")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_body(cell, fill=None, bold=False, center=False):
    cell.font = Font(name=FALLBACK_FONT, size=11, bold=bold, color="111827")
    if fill: cell.fill = fill
    cell.alignment = Alignment(
        horizontal=("center" if center else "left"),
        vertical="top", wrap_text=True
    )
    cell.border = BORDER

# ---------------------------------------------------------------------------
wb = Workbook()

# ===========================================================================
# Sheet 1: คณะทำงาน 17 คน
# ===========================================================================
ws1 = wb.active
ws1.title = "คณะทำงาน 17 คน"

ws1.merge_cells("A1:E1")
ws1["A1"] = ("คณะทำงานขับเคลื่อนศูนย์ความเป็นเลิศปัญญาประดิษฐ์ "
             "ด้านอุตสาหกรรมสร้างสรรค์ (COE AI – Creative Economy)")
style_title(ws1["A1"])
ws1.row_dimensions[1].height = 42

ws1.merge_cells("A2:E2")
ws1["A2"] = "รายชื่อและองค์ประกอบ จำนวน 17 ท่าน  |  จัดทำ พ.ค. 2569"
style_subtitle(ws1["A2"])
ws1.row_dimensions[2].height = 22

headers = ["ลำดับ", "ตำแหน่งในคณะทำงาน", "ชื่อ–สกุล / ผู้แทน", "สังกัด / หน่วยงาน", "บทบาท & ความเชี่ยวชาญ"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    style_header(c)
ws1.row_dimensions[4].height = 36

members = [
    # (ตำแหน่ง, ผู้แทน(เว้นให้กรอก), สังกัด, บทบาท)
    ("ประธานคณะทำงาน",
     "(โปรดระบุ)",
     "ปลัดกระทรวงดิจิทัลเพื่อเศรษฐกิจและสังคม / กระทรวง อว.",
     "กำหนดทิศทางเชิงยุทธศาสตร์ของ COE / กำกับนโยบายภาพรวม"),
    ("รองประธาน คนที่ 1",
     "(โปรดระบุ)",
     "ผู้อำนวยการสำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "เชื่อมโยงนโยบายเศรษฐกิจสร้างสรรค์ของประเทศ"),
    ("รองประธาน คนที่ 2",
     "(โปรดระบุ)",
     "ผู้อำนวยการศูนย์ AI แห่งชาติ / สวทช. / NECTEC",
     "ขับเคลื่อนเทคโนโลยี AI และโครงสร้างพื้นฐาน"),
    ("กรรมการ (ภาควิชาการ)",
     "(โปรดระบุ)",
     "ผู้แทนมหาวิทยาลัย — Computer Science / AI Research",
     "วิจัยพื้นฐาน Generative AI, LLM, Multimodal Models"),
    ("กรรมการ (ภาควิชาการ – ออกแบบ)",
     "(โปรดระบุ)",
     "ผู้แทนคณะออกแบบ / สถาปัตย์ / นิเทศศิลป์",
     "เชื่อม AI กับ Design Thinking & Creative Methodology"),
    ("กรรมการ (ภาพยนตร์ & แอนิเมชัน)",
     "(โปรดระบุ)",
     "สมาพันธ์สมาคมภาพยนตร์ฯ / TACGA",
     "ผู้ใช้ AI ใน VFX, Animation, Virtual Production"),
    ("กรรมการ (เพลง & เสียง)",
     "(โปรดระบุ)",
     "สมาคมการค้าผู้ประกอบธุรกิจบันเทิงไทย / ค่ายเพลง",
     "AI สำหรับการผลิตเพลง / Sound Design / Mastering"),
    ("กรรมการ (เกม & E-Sports)",
     "(โปรดระบุ)",
     "สมาคมอุตสาหกรรมซอฟต์แวร์เกมไทย (TGA)",
     "AI Game Dev / Procedural Content / NPC Behaviour"),
    ("กรรมการ (โฆษณา & คอนเทนต์ดิจิทัล)",
     "(โปรดระบุ)",
     "สมาคมโฆษณาแห่งประเทศไทย / DAAT",
     "Generative AI สำหรับ Campaign / Personalisation"),
    ("กรรมการ (แฟชั่น & งานฝีมือ)",
     "(โปรดระบุ)",
     "สภาอุตสาหกรรมแฟชั่น / สมาคมผู้ออกแบบเครื่องแต่งกาย",
     "AI Design Tools / Trend Forecasting / Soft Power"),
    ("กรรมการ (สถาปัตยกรรม & เมือง)",
     "(โปรดระบุ)",
     "สมาคมสถาปนิกสยาม / TCDC",
     "AI Generative Design, Urban Creative"),
    ("กรรมการ (ผู้ประกอบการ AI/Tech Startup)",
     "(โปรดระบุ)",
     "Thai AI Startup / สมาคม Thailand Tech Startup",
     "เชื่อมโยง AI Product สู่ Creative Industry"),
    ("กรรมการ (ทรัพย์สินทางปัญญา & กฎหมาย)",
     "(โปรดระบุ)",
     "กรมทรัพย์สินทางปัญญา / นักกฎหมาย AI Ethics",
     "นโยบายลิขสิทธิ์ AI, Data Governance, Ethics"),
    ("กรรมการ (ภาคการเงิน & ทุน)",
     "(โปรดระบุ)",
     "BOI / depa / ธนาคารพัฒนาวิสาหกิจ (SME D Bank)",
     "กลไกการลงทุน, ทุนวิจัย, สิทธิประโยชน์"),
    ("กรรมการ (ที่ปรึกษาผู้ทรงคุณวุฒิ)",
     "(โปรดระบุ)",
     "ผู้เชี่ยวชาญระดับนานาชาติด้าน AI x Creative",
     "ให้คำปรึกษาเชิงยุทธศาสตร์และ Benchmark โลก"),
    ("เลขานุการคณะทำงาน",
     "(โปรดระบุ)",
     "ผู้แทนหน่วยงานเจ้าภาพ COE",
     "บริหารจัดการการประชุม, ติดตามผล, รายงาน"),
    ("ผู้ช่วยเลขานุการ",
     "(โปรดระบุ)",
     "เจ้าหน้าที่อาวุโสฝ่ายยุทธศาสตร์",
     "สนับสนุนงานเลขานุการและประสานคณะอนุกรรมการ"),
]

assert len(members) == 17, f"ต้องเป็น 17 คน แต่ได้ {len(members)}"

start_row = 5
for i, (pos, name, org, role) in enumerate(members, 1):
    r = start_row + i - 1
    row_fill = ZEBRA_FILL if i % 2 == 0 else None
    bold = pos.startswith("ประธาน") or pos.startswith("รองประธาน") or "เลขานุการ" in pos
    style_body(ws1.cell(row=r, column=1, value=i), fill=row_fill, bold=True, center=True)
    style_body(ws1.cell(row=r, column=2, value=pos), fill=row_fill, bold=bold)
    style_body(ws1.cell(row=r, column=3, value=name), fill=row_fill)
    style_body(ws1.cell(row=r, column=4, value=org), fill=row_fill)
    style_body(ws1.cell(row=r, column=5, value=role), fill=row_fill)
    ws1.row_dimensions[r].height = 42

# column widths
for col, w in zip("ABCDE", [6, 32, 22, 38, 46]):
    ws1.column_dimensions[col].width = w

ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A5"

# ===========================================================================
# Sheet 2: เป้าหมาย / Output / Outcomes 1-3-5 ปี
# ===========================================================================
ws2 = wb.create_sheet("เป้าหมาย 1-3-5 ปี")

ws2.merge_cells("A1:E1")
ws2["A1"] = "เป้าหมาย Output / Outcomes ของ COE AI – Creative Economy  (ระยะ 1 ปี / 3 ปี / 5 ปี)"
style_title(ws2["A1"])
ws2.row_dimensions[1].height = 42

ws2.merge_cells("A2:E2")
ws2["A2"] = ("ปี 1 = Foundation (2569–2570)  |  ปี 3 = Scaling (2571–2572)  |  "
             "ปี 5 = Leadership (2573–2574)")
style_subtitle(ws2["A2"])
ws2.row_dimensions[2].height = 22

# Header row
hdr = ["มิติ (Pillar)",
       "ปี 1 — Foundation (Output)",
       "ปี 3 — Scaling (Outcome)",
       "ปี 5 — Leadership (Impact)",
       "ตัวชี้วัด (KPI)"]
for c, h in enumerate(hdr, 1):
    style_header(ws2.cell(row=4, column=c, value=h))
ws2.row_dimensions[4].height = 38

# สี Fill ต่อปี
y1f, y3f, y5f = Y1_FILL, Y3_FILL, Y5_FILL

pillars = [
    # (Pillar, Y1, Y3, Y5, KPI)
    ("1) วิจัย & พัฒนา AI\n(R&D)",
     "• ตั้งห้องปฏิบัติการ AI x Creative 1 แห่ง\n"
     "• งานวิจัย/Prototype 10 ชิ้น\n"
     "• ตีพิมพ์ระดับนานาชาติ 5 ฉบับ",
     "• Thai Creative Foundation Model (TH-GenAI) v1\n"
     "• Open Datasets อย่างน้อย 5 ชุด\n"
     "• สิทธิบัตร/ลิขสิทธิ์ AI 10 รายการ",
     "• TH-GenAI v3 ติด Top-5 ASEAN\n"
     "• ศูนย์วิจัยร่วมกับ 3 มหาวิทยาลัยระดับโลก\n"
     "• ผลงานวิจัยถูกอ้างอิงระดับ Tier-1",
     "จำนวนงานวิจัย, สิทธิบัตร, Citations, โมเดล Open Source"),

    ("2) กำลังคน & ทักษะ\n(Talent)",
     "• หลักสูตรอบรม AI x Creative 3 หลักสูตร\n"
     "• ผู้ผ่านอบรม 500 คน\n"
     "• เครือข่าย Mentor 30 คน",
     "• Reskill/Upskill 5,000 คน\n"
     "• ปริญญาตรี/โท สาขา AI–Creative\n"
     "• Talent Pool ทะเบียน 2,000 คน",
     "• 20,000 คนในระบบ\n"
     "• Thai Talent ทำงานเวทีโลก 200+ คน\n"
     "• กลายเป็น Hub การเรียน AI-Creative ของอาเซียน",
     "จำนวนผู้ผ่านอบรม, อัตราการได้งาน, รายได้เฉลี่ยเพิ่มขึ้น"),

    ("3) แพลตฟอร์ม & โครงสร้างพื้นฐาน\n(Infrastructure)",
     "• Sandbox AI Compute (GPU) ใช้งานร่วม\n"
     "• แพลตฟอร์มแบ่งปันข้อมูล (Data Hub) เปิดตัว\n"
     "• MOU พันธมิตร 10 องค์กร",
     "• Sovereign AI Compute รองรับ 50 องค์กร\n"
     "• Creative Data Hub 50 ชุดข้อมูล\n"
     "• API & Toolkits มาตรฐานเปิดใช้",
     "• ASEAN Creative AI Cloud Hub\n"
     "• เชื่อมต่อกับ Global Compute Network\n"
     "• Marketplace AI Tools ของไทยใช้ทั่วภูมิภาค",
     "GPU-hours ที่ใช้, จำนวนผู้ใช้, จำนวน Dataset"),

    ("4) อุตสาหกรรม & ผู้ประกอบการ\n(Industry)",
     "• 100 SME/Studio นำ AI ไปใช้จริง\n"
     "• Pilot Projects 20 โครงการ\n"
     "• เปิดตัว Showcase 1 ครั้ง/ปี",
     "• 1,000 บริษัทใช้ AI Creative Stack\n"
     "• Creative AI Startup ใหม่ 50 ราย\n"
     "• ส่งออกคอนเทนต์ AI-assisted +30%",
     "• 5,000+ บริษัทในระบบ\n"
     "• Thai IP / Content ส่งออกระดับโลก\n"
     "• สร้าง Unicorn ด้าน Creative AI 1–2 ราย",
     "มูลค่าเพิ่มต่อ GDP, จำนวนบริษัท, มูลค่าส่งออก"),

    ("5) นโยบาย & ธรรมาภิบาล\n(Policy & Ethics)",
     "• ร่างกรอบจริยธรรม AI x Creative\n"
     "• คู่มือลิขสิทธิ์ AI Content\n"
     "• 1 Public Hearing",
     "• ประกาศใช้แนวปฏิบัติระดับชาติ\n"
     "• กลไก IP Registry สำหรับ AI Works\n"
     "• Sandbox นโยบาย",
     "• ไทยเป็นผู้นำมาตรฐาน AI-Creative ใน ASEAN\n"
     "• ส่งออกโมเดล Policy ให้ประเทศอื่นใช้อ้างอิง",
     "จำนวนนโยบายที่ออก, การยอมรับสากล"),

    ("6) ความร่วมมือระดับโลก\n(Global Partnership)",
     "• MOU ต่างประเทศ 3 ประเทศ\n"
     "• เข้าร่วมเวทีโลก 5 เวที",
     "• Joint Lab กับ 3 ประเทศ\n"
     "• Thailand Pavilion ในงาน Top-3 ของโลก\n"
     "• ทุนวิจัยร่วม ≥ 100 ล้านบาท",
     "• Thailand เป็นเจ้าภาพ AI x Creative World Summit\n"
     "• Network 20+ ประเทศ\n"
     "• Soft Power Index ของไทยเพิ่มขึ้น",
     "จำนวน MOU, ทุนต่างประเทศ, อันดับ Soft Power"),

    ("7) ผลกระทบเชิงเศรษฐกิจ\n(Economic Impact)",
     "• ประเมิน Baseline GDP สาขาสร้างสรรค์\n"
     "• Roadmap Economic Model",
     "• เพิ่ม GVA อุตสาหกรรมสร้างสรรค์ +10%\n"
     "• การจ้างงานใหม่ ≥ 10,000 ตำแหน่ง",
     "• เพิ่ม GVA +25–30% เทียบ Baseline\n"
     "• ดันสัดส่วน Creative Economy ต่อ GDP ≥ 12%\n"
     "• สร้างงาน 50,000+ ตำแหน่ง",
     "GDP สาขาสร้างสรรค์, การจ้างงาน, มูลค่าส่งออก"),
]

start = 5
for i, (p, y1, y3, y5, kpi) in enumerate(pillars):
    r = start + i
    style_body(ws2.cell(row=r, column=1, value=p), bold=True)
    style_body(ws2.cell(row=r, column=2, value=y1), fill=y1f)
    style_body(ws2.cell(row=r, column=3, value=y3), fill=y3f)
    style_body(ws2.cell(row=r, column=4, value=y5), fill=y5f)
    style_body(ws2.cell(row=r, column=5, value=kpi))
    ws2.row_dimensions[r].height = 105

# Summary box
sum_row = start + len(pillars) + 1
ws2.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)
c = ws2.cell(row=sum_row, column=1,
             value="วิสัยทัศน์ 5 ปี:  ทำให้ประเทศไทยเป็น “ASEAN Hub of AI-powered Creative Economy” "
                   "ที่ผลักดัน Soft Power ไทยสู่เวทีโลก ด้วยกำลังคน เทคโนโลยี และระบบนิเวศที่พร้อม")
c.font = Font(name=FALLBACK_FONT, size=12, bold=True, color="FFFFFF")
c.fill = PRIMARY_FILL
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
c.border = BORDER
ws2.row_dimensions[sum_row].height = 50

# widths
for col, w in zip("ABCDE", [22, 36, 36, 36, 28]):
    ws2.column_dimensions[col].width = w

ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

# ---------------------------------------------------------------------------
out = "COE_AI_Creative_Economy.xlsx"
wb.save(out)
print(f"saved: {out}")
