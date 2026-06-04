"""สร้าง Excel v4 — COE AI Creative Economy (Decision Document)
เป้าหมาย: เทียบ 2 ทีม + Output/Outcome อย่างเดียว

4 Sheets:
  1. ชุด 1 — คณะทำงาน 17 คน (CEA ประธาน)
  2. ชุด 2 — คณะทำงาน 17 คน (AICAT ประธาน)
  3. เปรียบเทียบ ชุด 1 vs ชุด 2
  4. แผน Output/Outcome 1-3-5 ปี (ตัด Soft Power)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
thin = Side(border_style="thin", color="B0B8C1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Colour palette
C_NAVY   = "1F4E78"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_GREEN  = "E2EFDA"
C_YELLOW = "FFF2CC"
C_ORANGE = "FCE4D6"
C_GRAY   = "F2F7FA"
C_STRIPE = "F8F9FA"
C_RED_SOFT = "FDE8E8"
C_PURPLE = "EDE7F6"
C_TEAL   = "D1FAE5"
C_DIFF   = "FFE4B5"  # highlight diff

def fill(hex_color): return PatternFill("solid", start_color=hex_color)
def font(size=11, bold=False, color="111827", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_cell(ws, row, col, value, bg=C_NAVY, fg="FFFFFF", size=11, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h=h, v="center")
    c.border = BORDER

def data_cell(ws, row, col, value, bg=None, bold=False, h="left", size=11, italic=False, color="111827"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if bg: c.fill = fill(bg)
    c.alignment = align(h=h)
    c.border = BORDER

def merge_title(ws, row, col_start, col_end, value, bg=C_NAVY, fg="FFFFFF", size=14, height=40):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font = Font(name=FONT, size=size, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h="center", v="center")
    c.border = BORDER
    ws.row_dimensions[row].height = height

BLOCK_COLORS = {
    "A": C_PURPLE,
    "B": C_GREEN,
    "C": C_YELLOW,
    "D": C_ORANGE,
}

def build_roster_sheet(ws, title, subtitle, members, sheet_color=C_NAVY):
    """สร้าง sheet roster มาตรฐาน"""
    ws.sheet_view.showGridLines = False
    merge_title(ws, 1, 1, 6, title, bg=sheet_color)
    merge_title(ws, 2, 1, 6, subtitle, bg=C_BLUE, size=10, height=22)

    headers = ["ลำดับ", "บล็อก", "ตำแหน่ง", "ชื่อ–สกุล / ผู้แทน", "สังกัด / หน่วยงาน", "บทบาท"]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 3, col, h, bg=C_LBLUE, fg="1F2937")
    ws.row_dimensions[3].height = 30

    for i, (blk, pos, name, org, role) in enumerate(members, 1):
        r = 3 + i
        bg = BLOCK_COLORS[blk]
        data_cell(ws, r, 1, i, bg=bg, bold=True, h="center")
        data_cell(ws, r, 2, blk, bg=bg, bold=True, h="center")
        data_cell(ws, r, 3, pos, bg=bg, bold=(blk in ("A","D")))
        data_cell(ws, r, 4, name, bg=bg)
        data_cell(ws, r, 5, org, bg=bg)
        data_cell(ws, r, 6, role, bg=bg)
        ws.row_dimensions[r].height = 48

    for col, w in zip(range(1, 7), [5, 6, 26, 22, 36, 38]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

wb = Workbook()

# ══════════════════════════════════════════════════════
# SHEET 1: ชุด 1 — คณะทำงาน 17 คน (CEA ประธาน)
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "ชุด 1 — CEA ประธาน"

members_v1 = [
    # Block A — ผู้นำ (3)
    ("A","ประธานคณะทำงาน",
     "ดร.ชาคริต พิชญางกูร",
     "สำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "กำหนดทิศทาง / Champion AI x Creative Economy"),
    ("A","รองประธานคนที่ 1",
     "นายสุธัช เจริญผล",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานภาคเอกชน / ขับเคลื่อนงานปฏิบัติ"),
    ("A","รองประธานคนที่ 2",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานส่งเสริมเศรษฐกิจดิจิทัล (depa)",
     "เชื่อม Digital Economy / Infrastructure"),

    # Block B — กรรมการรัฐ (5)
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมวัฒนธรรม กระทรวงวัฒนธรรม",
     "เศรษฐกิจวัฒนธรรม + มรดกทางวัฒนธรรม"),
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมทรัพย์สินทางปัญญา กระทรวงพาณิชย์",
     "AI Copyright / IP Framework"),
    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานนวัตกรรมแห่งชาติ (NIA)",
     "Grant + Startup Funding"),
    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สถาบันข้อมูลขนาดใหญ่ (BDI)",
     "Thai Creative AI Dataset"),
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมการค้าระหว่างประเทศ (DITP)",
     "Export + Bangkok Content Market"),

    # Block C — ผู้ทรงคุณวุฒิ (6)
    ("C","ผู้ทรงคุณวุฒิ (1)",
     "(นายก / ผู้แทน)",
     "สมาคมโฆษณาดิจิทัล (DAAT)",
     "Generative AI for Digital Advertising"),
    ("C","ผู้ทรงคุณวุฒิ (2)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้ประกอบการแอนิเมชันฯ (TACGA)",
     "AI in Animation / VFX"),
    ("C","ผู้ทรงคุณวุฒิ (3)",
     "(ผู้แทน)",
     "สภาอุตสาหกรรมแห่งประเทศไทย (FTI) กลุ่มดิจิทัล/Creative",
     "Voice ภาคอุตสาหกรรมสร้างสรรค์รวม"),
    ("C","ผู้ทรงคุณวุฒิ (4)",
     "อ.สมชาย (โปรดระบุนามสกุล)",
     "ผู้เชี่ยวชาญด้านกฎหมาย IP และ AI",
     "AI Copyright Law / IP Framework"),
    ("C","ผู้ทรงคุณวุฒิ (5)",
     "(ผู้อำนวยการ / ผู้แทน)",
     "MACI คณะดุริยางคศาสตร์ ม.ศิลปากร",
     "Music + AI / Academic-Industry Bridge"),
    ("C","ผู้ทรงคุณวุฒิ (6)",
     "(นายก / ผู้แทน)",
     "สมาคม Content Creator",
     "Creator Economy / UGC + AI Tools"),

    # Block D — เลขานุการ (3)
    ("D","กรรมการและเลขานุการ",
     "(ผู้แทนระดับปฏิบัติ CEA)",
     "สำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "บริหารวาระประชุม / ติดตามผล"),
    ("D","กรรมการและเลขานุการร่วม",
     "(ผู้แทนระดับปฏิบัติ สดช.)",
     "สำนักงานคณะกรรมการดิจิทัลฯ (สดช.)",
     "รายงานกลับ คกก.แม่"),
    ("D","กรรมการและผู้ช่วยเลขานุการ",
     "(ผู้แทน AICAT)",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานภาคเอกชน / กิจกรรม COE"),
]
assert len(members_v1) == 17

build_roster_sheet(ws1,
    "ชุด 1 — คณะทำงาน 17 คน (CEA เป็นประธาน)",
    "Government-led Balanced | ดร.ชาคริต ประธาน + AICAT รองประธาน 1 + depa รองประธาน 2 | เลขา 3 ท่าน | ไม่มีที่ปรึกษา",
    members_v1, sheet_color="1F4E78")

# ══════════════════════════════════════════════════════
# SHEET 2: ชุด 2 — คณะทำงาน 17 คน (AICAT ประธาน)
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("ชุด 2 — AICAT ประธาน")

members_v2 = [
    # Block A — ผู้นำ (2)
    ("A","ประธานคณะทำงาน",
     "นายสุธัช เจริญผล",
     "นายกสมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "กำหนดทิศทาง / เจ้าภาพประสานทุกฝ่าย"),
    ("A","รองประธานคณะทำงาน",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานส่งเสริมเศรษฐกิจดิจิทัล (depa)",
     "เชื่อม Digital Economy / ทำหน้าที่แทนประธาน"),

    # Block B — กรรมการรัฐ (6)
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมวัฒนธรรม กระทรวงวัฒนธรรม",
     "เศรษฐกิจวัฒนธรรม + มรดกทางวัฒนธรรม"),
    ("B","กรรมการ",
     "ดร.ชาคริต พิชญางกูร",
     "สำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "Champion AI x Creative / Mandate รองนายกฯ"),
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมทรัพย์สินทางปัญญา กระทรวงพาณิชย์",
     "AI Copyright / IP Framework"),
    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานนวัตกรรมแห่งชาติ (NIA)",
     "Grant + Startup Funding"),
    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สถาบันข้อมูลขนาดใหญ่ (BDI)",
     "Thai Creative AI Dataset"),
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมการค้าระหว่างประเทศ (DITP)",
     "Export + Bangkok Content Market"),

    # Block C — ผู้ทรงคุณวุฒิ (7)
    ("C","ผู้ทรงคุณวุฒิ (1)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้กำกับภาพยนตร์ไทย",
     "อุตสาหกรรมภาพยนตร์ + AI Film Production"),
    ("C","ผู้ทรงคุณวุฒิ (2)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้ประกอบการแอนิเมชันฯ (TACGA)",
     "AI in Animation / VFX"),
    ("C","ผู้ทรงคุณวุฒิ (3)",
     "(นายก / ผู้แทน)",
     "สมาคมสมาพันธ์โอเพนซอร์สแห่งประเทศไทย",
     "Open Source AI + Foundation Model ไทย"),
    ("C","ผู้ทรงคุณวุฒิ (4)",
     "(โปรดระบุชื่อ–นามสกุล)",
     "ผู้เชี่ยวชาญด้าน CGI และวิชวลเอฟเฟกต์",
     "AI ใน CGI / VFX / Production Pipeline"),
    ("C","ผู้ทรงคุณวุฒิ (5)",
     "(นายก / ผู้แทน)",
     "สมาคมโฆษณาดิจิทัล (DAAT)",
     "Generative AI for Digital Advertising"),
    ("C","ผู้ทรงคุณวุฒิ (6)",
     "(นายก / ผู้แทน)",
     "สมาคม Content Creator",
     "Creator Economy / UGC + AI Tools"),
    ("C","ผู้ทรงคุณวุฒิ (7)",
     "อ.สมชาย (โปรดระบุนามสกุล)",
     "ผู้เชี่ยวชาญด้านกฎหมาย IP และ AI",
     "AI Copyright Law / IP Framework"),

    # Block D — เลขานุการ (2)
    ("D","กรรมการและเลขานุการ",
     "(ผู้แทนระดับปฏิบัติ สดช.)",
     "สำนักงานคณะกรรมการดิจิทัลฯ (สดช.)",
     "รายงานกลับ คกก.แม่ / บริหารประชุม"),
    ("D","กรรมการและเลขานุการร่วม",
     "(ผู้แทน AICAT)",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานเอกชน / กิจกรรม COE"),
]
assert len(members_v2) == 17

build_roster_sheet(ws2,
    "ชุด 2 — คณะทำงาน 17 คน (AICAT เป็นประธาน)",
    "AICAT-led Industry-deep | นายสุธัช เจริญผล ประธาน + depa รองประธาน | CEA = กรรมการ | เลขา 2 + ที่ปรึกษา 4 (นอก 17)",
    members_v2, sheet_color="065F46")

# Add ที่ปรึกษา section at bottom of Sheet 2
adv_start = 4 + len(members_v2) + 2
merge_title(ws2, adv_start, 1, 6,
    "ที่ปรึกษาคณะทำงาน 4 ท่าน (เพิ่มเติม — นอก 17 ที่นั่ง)",
    bg="065F46", size=12, height=30)

adv_header_row = adv_start + 1
for col, h in enumerate(["ลำดับ", "ตำแหน่ง", "ชื่อ / ผู้แทน", "สังกัด", "ความเชี่ยวชาญ", ""], 1):
    if h:
        header_cell(ws2, adv_header_row, col, h, bg=C_LBLUE, fg="1F2937")
    else:
        header_cell(ws2, adv_header_row, col, "", bg=C_LBLUE)
ws2.row_dimensions[adv_header_row].height = 28

advisors = [
    ("ป.1", "ที่ปรึกษา", "(ผู้บริหาร / ผู้แทน)", "WorkPoint Entertainment", "Mass Content / Distribution"),
    ("ป.2", "ที่ปรึกษา", "(โปรดระบุชื่อ)", "ผู้เชี่ยวชาญ Film Production Technology", "AI Production Pipeline"),
    ("ป.3", "ที่ปรึกษา", "(โปรดระบุชื่อ)", "นักวิจัย / อาจารย์ด้าน AI", "Academic AI Research"),
    ("ป.4", "ที่ปรึกษา", "(ผู้บริหาร / ผู้แทน)", "Bitkub Online", "Digital Asset / NFT for Creative"),
]
for i, (no, pos, name, org, exp) in enumerate(advisors, 1):
    r = adv_header_row + i
    bg = C_TEAL if i % 2 == 1 else None
    data_cell(ws2, r, 1, no, bg=bg, bold=True, h="center")
    data_cell(ws2, r, 2, pos, bg=bg, bold=True)
    data_cell(ws2, r, 3, name, bg=bg)
    data_cell(ws2, r, 4, org, bg=bg)
    ws2.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    data_cell(ws2, r, 5, exp, bg=bg)
    ws2.row_dimensions[r].height = 32

# ══════════════════════════════════════════════════════
# SHEET 3: เปรียบเทียบ ชุด 1 vs ชุด 2
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("เปรียบเทียบ ชุด 1 vs 2")
ws3.sheet_view.showGridLines = False

merge_title(ws3, 1, 1, 4,
    "เปรียบเทียบโครงสร้าง — ชุด 1 vs ชุด 2")
merge_title(ws3, 2, 1, 4,
    "ทั้ง 2 ชุด: 17 ที่นั่งเท่ากัน — แต่ Strategic Posture, ประธาน, และ Composition ต่างกัน",
    bg=C_BLUE, size=10, height=22)

# Section: Summary
sum_r = 3
for col, h in enumerate(["มิติ", "ชุด 1 (CEA ประธาน)", "ชุด 2 (AICAT ประธาน)", "Note"], 1):
    header_cell(ws3, sum_r, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[sum_r].height = 32

summary_rows = [
    ("Strategic Posture", "Government-led, Balanced", "AICAT-led, Industry-deep", ""),
    ("ประธาน", "ดร.ชาคริต — ผอ.CEA", "นายสุธัช เจริญผล — นายก AICAT", "⚠ จุดต่างสำคัญสุด"),
    ("รองประธาน", "2 คน (AICAT + depa)", "1 คน (depa)", "ชุด 2 ลด 1"),
    ("Block A (ผู้นำ)", "3 คน", "2 คน", ""),
    ("Block B (กรรมการรัฐ)", "5 หน่วย", "6 หน่วย (+CEA)", "CEA ลงเป็นกรรมการในชุด 2"),
    ("Block C (ผู้ทรงคุณวุฒิ)", "6 คน", "7 คน", "ชุด 2 industry deep กว่า"),
    ("Block D (เลขานุการ)", "3 คน", "2 คน", "ชุด 2 ตัด CEA staff"),
    ("ที่ปรึกษา (นอก 17)", "ไม่มี", "4 ท่าน", "ชุด 2 มี mechanism เพิ่ม"),
    ("Political Risk", "ต่ำ (ตามขนบ)", "กลาง-สูง (ต้อง pre-sync ดร.ชาคริต)", "⚠"),
    ("ความคล่องตัว", "ติด protocol รัฐ", "คล่อง — AICAT decide เร็ว", ""),
    ("Industry coverage", "Animation+Music+Law", "+Film+Open Source+CGI", "ชุด 2 ครอบมากกว่า"),
]
for i, row in enumerate(summary_rows):
    r = sum_r + 1 + i
    bg = C_STRIPE if i % 2 == 0 else None
    is_critical = row[3].startswith("⚠")
    if is_critical:
        bg = C_RED_SOFT
    data_cell(ws3, r, 1, row[0], bg=bg, bold=True)
    data_cell(ws3, r, 2, row[1], bg=bg)
    data_cell(ws3, r, 3, row[2], bg=bg)
    data_cell(ws3, r, 4, row[3], bg=bg, italic=True, color="B91C1C" if is_critical else "374151")
    ws3.row_dimensions[r].height = 28

# Section: Block-by-Block Roster
block_start = sum_r + 1 + len(summary_rows) + 2

# Block A
merge_title(ws3, block_start, 1, 4, "Block A — ผู้นำ", bg=C_PURPLE, fg="111827", size=12, height=26)
ba_h = block_start + 1
for col, h in enumerate(["ตำแหน่ง", "ชุด 1", "ชุด 2", "Diff"], 1):
    header_cell(ws3, ba_h, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[ba_h].height = 26
ba_rows = [
    ("ประธาน", "ดร.ชาคริต — CEA", "นายสุธัช เจริญผล — AICAT", "⚠ เปลี่ยน"),
    ("รองประธาน 1", "นายสุธัช เจริญผล — AICAT", "ผอ. — depa", "เปลี่ยน"),
    ("รองประธาน 2", "ผอ. — depa", "— (ไม่มี)", "ตัด"),
]
for i, row in enumerate(ba_rows):
    r = ba_h + 1 + i
    bg = C_RED_SOFT if "⚠" in row[3] else (C_STRIPE if i % 2 == 0 else None)
    data_cell(ws3, r, 1, row[0], bg=bg, bold=True)
    data_cell(ws3, r, 2, row[1], bg=bg)
    data_cell(ws3, r, 3, row[2], bg=bg)
    data_cell(ws3, r, 4, row[3], bg=bg, italic=True)
    ws3.row_dimensions[r].height = 28

# Block B
bb_start = ba_h + 1 + len(ba_rows) + 1
merge_title(ws3, bb_start, 1, 4, "Block B — กรรมการรัฐ", bg=C_GREEN, fg="111827", size=12, height=26)
bb_h = bb_start + 1
for col, h in enumerate(["#", "ชุด 1 (5 หน่วย)", "ชุด 2 (6 หน่วย)", "Diff"], 1):
    header_cell(ws3, bb_h, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[bb_h].height = 26
bb_rows = [
    ("1", "กรมส่งเสริมวัฒนธรรม", "กรมส่งเสริมวัฒนธรรม", "✓"),
    ("2", "กรมทรัพย์สินทางปัญญา", "CEA (ดร.ชาคริต)", "⚠ ใหม่"),
    ("3", "NIA", "กรมทรัพย์สินทางปัญญา", "shift"),
    ("4", "BDI", "NIA", "shift"),
    ("5", "DITP", "BDI", "shift"),
    ("6", "—", "DITP", "+เพิ่ม"),
]
for i, row in enumerate(bb_rows):
    r = bb_h + 1 + i
    bg = C_RED_SOFT if "⚠" in row[3] else (C_STRIPE if i % 2 == 0 else None)
    data_cell(ws3, r, 1, row[0], bg=bg, bold=True, h="center")
    data_cell(ws3, r, 2, row[1], bg=bg)
    data_cell(ws3, r, 3, row[2], bg=bg)
    data_cell(ws3, r, 4, row[3], bg=bg, italic=True)
    ws3.row_dimensions[r].height = 26

# Block C
bc_start = bb_h + 1 + len(bb_rows) + 1
merge_title(ws3, bc_start, 1, 4, "Block C — ผู้ทรงคุณวุฒิ", bg=C_YELLOW, fg="111827", size=12, height=26)
bc_h = bc_start + 1
for col, h in enumerate(["#", "ชุด 1 (6 คน)", "ชุด 2 (7 คน)", "Diff"], 1):
    header_cell(ws3, bc_h, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[bc_h].height = 26
bc_rows = [
    ("1", "DAAT", "สมาคมผู้กำกับภาพยนตร์ไทย", "+ใหม่"),
    ("2", "TACGA", "TACGA", "✓"),
    ("3", "FTI สภาอุตสาหกรรม", "สมาคมสมาพันธ์โอเพนซอร์ส", "เปลี่ยน"),
    ("4", "อ.สมชาย (Law)", "ผู้เชี่ยวชาญ CGI", "เปลี่ยน"),
    ("5", "MACI ม.ศิลปากร", "DAAT", "เปลี่ยน"),
    ("6", "สมาคม Content Creator", "สมาคม Content Creator", "✓"),
    ("7", "—", "อ.สมชาย (Law)", "+อ.สมชาย กลับ"),
]
for i, row in enumerate(bc_rows):
    r = bc_h + 1 + i
    bg = C_STRIPE if i % 2 == 0 else None
    data_cell(ws3, r, 1, row[0], bg=bg, bold=True, h="center")
    data_cell(ws3, r, 2, row[1], bg=bg)
    data_cell(ws3, r, 3, row[2], bg=bg)
    data_cell(ws3, r, 4, row[3], bg=bg, italic=True)
    ws3.row_dimensions[r].height = 26

# Block D
bd_start = bc_h + 1 + len(bc_rows) + 1
merge_title(ws3, bd_start, 1, 4, "Block D — เลขานุการ", bg=C_ORANGE, fg="111827", size=12, height=26)
bd_h = bd_start + 1
for col, h in enumerate(["#", "ชุด 1 (3 คน)", "ชุด 2 (2 คน)", "Diff"], 1):
    header_cell(ws3, bd_h, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[bd_h].height = 26
bd_rows = [
    ("1", "CEA staff", "สดช.", "ตัด CEA staff"),
    ("2", "สดช.", "AICAT", "shift"),
    ("3", "AICAT", "— (ไม่มี)", "ตัด"),
]
for i, row in enumerate(bd_rows):
    r = bd_h + 1 + i
    bg = C_STRIPE if i % 2 == 0 else None
    data_cell(ws3, r, 1, row[0], bg=bg, bold=True, h="center")
    data_cell(ws3, r, 2, row[1], bg=bg)
    data_cell(ws3, r, 3, row[2], bg=bg)
    data_cell(ws3, r, 4, row[3], bg=bg, italic=True)
    ws3.row_dimensions[r].height = 26

# Decision callout
dec_start = bd_h + 1 + len(bd_rows) + 2
ws3.merge_cells(start_row=dec_start, start_column=1, end_row=dec_start, end_column=4)
dc = ws3.cell(row=dec_start, column=1,
    value="🎯 คำแนะนำเชิงกลยุทธ์")
dc.font = Font(name=FONT, size=13, bold=True, color="FFFFFF")
dc.fill = fill(C_NAVY)
dc.alignment = align(h="center", v="center")
dc.border = BORDER
ws3.row_dimensions[dec_start].height = 32

rec_rows = [
    ("เป้าหมาย", "เลือกชุด", "เหตุผล"),
    ("ลด political risk + ส่ง สดช. ผ่านง่าย", "ชุด 1", "ตามขนบ คกก.แม่ approve ง่ายกว่า"),
    ("AICAT มี ownership ชัด + ขับเคลื่อนเร็ว", "ชุด 2", "AICAT-led + industry network พร้อม"),
    ("Best of both worlds", "ชุด 2 + Pre-sync ดร.ชาคริต", "ก่อน lock ชุด 2 ต้องคุย ดร.ชาคริต ให้ยอมรับเป็นกรรมการ"),
]
for i, row in enumerate(rec_rows):
    r = dec_start + 1 + i
    if i == 0:
        for col, val in enumerate(row, 1):
            header_cell(ws3, r, col, val, bg=C_LBLUE, fg="1F2937")
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws3.row_dimensions[r].height = 26
    else:
        bg = C_STRIPE if i % 2 == 0 else None
        data_cell(ws3, r, 1, row[0], bg=bg)
        data_cell(ws3, r, 2, row[1], bg=bg, bold=True, h="center")
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        data_cell(ws3, r, 3, row[2], bg=bg, italic=True)
        ws3.row_dimensions[r].height = 30

for col, w in zip(range(1, 5), [22, 34, 34, 20]):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 4: แผน Output/Outcome 1-3-5 ปี (ตัด Soft Power)
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet("แผน Output-Outcome 1-3-5 ปี")
ws4.sheet_view.showGridLines = False

merge_title(ws4, 1, 1, 5,
    "เป้าหมาย Output / Outcome ของ COE AI Creative Economy  |  ระยะ 1 ปี / 3 ปี / 5 ปี")
merge_title(ws4, 2, 1, 5,
    "ปี 1 = Foundation 2569–70  |  ปี 3 = Scale 2571–72  |  ปี 5 = Leadership 2573–74",
    bg=C_BLUE, size=10, height=22)

h4 = ["มิติ (Pillar)", "ปี 1 — Foundation\n(Output)", "ปี 3 — Scale\n(Outcome)", "ปี 5 — Leadership\n(Impact)", "KPI / ตัวชี้วัด"]
for col, h in enumerate(h4, 1):
    header_cell(ws4, 3, col, h, bg=C_LBLUE, fg="1F2937")
ws4.row_dimensions[3].height = 42

pillars = [
    ("1) Hub &\nโครงสร้างพื้นฐาน",
     "• ตั้ง CoE Hub ร่วมกับ CEA/TCDC\n• เครือข่ายพันธมิตร MOU ≥ 10 องค์กร\n• แพลตฟอร์มกลาง AI Creative",
     "• ขยาย Hub ครอบคลุม 5 ภูมิภาค\n• เครือข่าย ASEAN ≥ 5 ประเทศ",
     "• ไทยเป็น AI Creative Hub ของ ASEAN\n• Hub เชื่อมต่อกับ Global AI Creative Network",
     "จำนวน MOU / จำนวน Hub / Active Users"),

    ("2) Talent &\nการศึกษา",
     "• เปิดหลักสูตร AI Creative Production ≥ 2 หลักสูตร\n• ผู้ผ่านอบรม ≥ 500 คน",
     "• Talent Pool ≥ 1,000 คน (ผ่านรับรองมาตรฐาน)\n• หลักสูตรขยายสู่ ≥ 10 สถาบัน",
     "• Talent Pipeline ป้อนอุตสาหกรรมต่อเนื่อง\n• Thai AI Creative Talent ทำงานในเวทีโลก",
     "จำนวนผู้ผ่านอบรม / อัตราการจ้างงาน"),

    ("3) Sandbox &\nนำร่อง",
     "• Sandbox 3–5 โปรเจกต์\n  (Animation / Music / Content / Ads)\n• สร้าง Use Case + Baseline",
     "• Scale จาก Pilot สู่ Commercial ≥ 10 โปรเจกต์\n• AI Creative Startup ≥ 20 ราย",
     "• AI Creative Startup/SME ที่ scale ในตลาด\n• Unicorn ด้าน Creative AI 1–2 ราย",
     "จำนวน Startup / มูลค่าเชิงพาณิชย์"),

    ("4) Thai Creative\nAI Dataset",
     "• Thai Creative AI Dataset Phase 1\n  (ภาษา วัฒนธรรม ทรัพยากรสร้างสรรค์ไทย)",
     "• Thai Creative AI Dataset Phase 2–3\n  (ขยาย sector + open access ภาคอุตสาหกรรม)",
     "• Dataset เป็น ASEAN Creative AI Benchmark\n• Thai AI Tools/IP ใช้งานใน ≥ 5 ประเทศ",
     "จำนวน Dataset / ประเทศที่ใช้งาน"),

    ("5) มาตรฐาน\n& กฎหมาย IP",
     "• กรอบนโยบาย AI Copyright & IP (เบื้องต้น)\n• แนวปฏิบัติ AI Creative Content ฉบับไทย",
     "• ประกาศใช้มาตรฐานวิชาชีพ AI Creative ระดับชาติ\n• จรรยาบรรณ AI Creative ฉบับไทย (official)",
     "• ไทยเป็นผู้นำมาตรฐาน AI Creative ใน ASEAN\n• ส่งออกโมเดลนโยบายให้ประเทศอื่นอ้างอิง",
     "จำนวนนโยบายที่ประกาศใช้ / การยอมรับสากล"),

    ("6) ต้นทุนผลิต\n& ผลิตภาพ",
     "• Baseline ต้นทุนผลิตคอนเทนต์ปัจจุบัน\n• เครื่องมือ AI ที่ผู้ประกอบการเข้าถึงได้",
     "• ต้นทุนผลิตคอนเทนต์ลดลง ≥ 30%\n  (เทียบ Baseline ปีที่ 1)",
     "• AI Adoption ≥ 80% ของผู้ประกอบการ Creative\n• ผลิตภาพเพิ่มขึ้นอย่างมีนัยสำคัญ",
     "% ต้นทุนลด / % AI Adoption"),

    # Pillar 7 — Soft Power → "เวทีระดับโลก / อิทธิพลทางวัฒนธรรม"
    ("7) Export &\nเวทีระดับโลก",
     "• ศึกษาตลาด ASEAN + จัดทำ Export Roadmap\n• เข้าร่วมงาน Bangkok International Content Market",
     "• เริ่ม export AI Creative Content ไทยสู่ ASEAN\n  ผ่าน DITP + Bangkok Content Market",
     "• Thai AI Creative Content ในเวทีระดับโลก\n  (รางวัล/ยอดชมระดับนานาชาติ)\n• อิทธิพลทางวัฒนธรรมไทยขยายตัวในเวทีโลก",
     "มูลค่าส่งออก / รางวัลนานาชาติ"),

    ("8) ผลกระทบ\nเชิงเศรษฐกิจ\n(KPI หลัก)",
     "• Baseline: Creative Economy = 8.78% ของ GDP\n  (CEA รายงาน พ.ค. 2569)",
     "• เพิ่มการจ้างงานสาขาสร้างสรรค์ ≥ 10,000 ตำแหน่ง\n• Creative AI Startup สร้างรายได้วัดได้",
     "• Creative Economy → 12%+ ของ GDP\n  โดย AI เป็นตัวขับเคลื่อนหลัก\n• ผู้ประกอบการใช้ AI ≥ 80% ของ sector",
     "% GDP / % AI Adoption\n(KPI หลัก 2 ตัว)"),
]

start4 = 4
for i, (pillar, y1, y3, y5, kpi) in enumerate(pillars):
    r = start4 + i
    is_last = (i == len(pillars) - 1)
    pillar_bg = C_RED_SOFT if is_last else None
    data_cell(ws4, r, 1, pillar, bg=pillar_bg, bold=is_last)
    data_cell(ws4, r, 2, y1, bg=C_GREEN)
    data_cell(ws4, r, 3, y3, bg=C_YELLOW)
    data_cell(ws4, r, 4, y5, bg=C_ORANGE)
    data_cell(ws4, r, 5, kpi, bold=is_last)
    ws4.row_dimensions[r].height = 95 if i < len(pillars)-1 else 115

vr = start4 + len(pillars) + 1
ws4.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=5)
vc = ws4.cell(row=vr, column=1,
    value='วิสัยทัศน์ 5 ปี: ทำให้ประเทศไทยเป็น "ASEAN Hub of AI-powered Creative Economy" '
          '— ผลักดันเศรษฐกิจวัฒนธรรมและอุตสาหกรรมสร้างสรรค์ไทยสู่เวทีโลก '
          'ด้วยกำลังคน เทคโนโลยี IP และระบบนิเวศที่พร้อม')
vc.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
vc.fill = fill(C_NAVY)
vc.alignment = align(h="center", v="center")
vc.border = BORDER
ws4.row_dimensions[vr].height = 50

for col, w in zip(range(1, 6), [18, 34, 34, 34, 26]):
    ws4.column_dimensions[get_column_letter(col)].width = w
ws4.freeze_panes = "B4"

# ══════════════════════════════════════════════════════
out = "COE_AI_Creative_Economy_v4.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
print(f"   Sheets: {wb.sheetnames}")

# Verify no Soft Power
import re
sp = 0
for sn in wb.sheetnames:
    for row in wb[sn].iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and re.search(r'soft\s*power', cell.value, re.IGNORECASE):
                sp += 1
                print(f"   ⚠ Soft Power found: {sn}: {cell.value[:60]}")
print(f"   ✅ Soft Power check: {'passed (0 found)' if sp == 0 else f'FAILED ({sp} found)'}")
