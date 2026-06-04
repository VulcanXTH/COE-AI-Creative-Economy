"""สร้าง Excel v3 — COE AI Creative Economy (รวมการตัดสินใจล่าสุดทั้งหมด)
อ้างอิงโครงสร้างจาก v2.xlsx แต่:
  - คณะทำงาน → ชุด 2 final (AICAT ประธาน + DITP + ที่ปรึกษา)
  - Output/Outcome → ตัด "Soft Power" ใช้ "เศรษฐกิจวัฒนธรรม / อิทธิพลทางวัฒนธรรม"
  - เพิ่ม Sheet ที่ปรึกษา + งบประมาณ

5 Sheets:
  1. คณะทำงาน 17 คน (Final v2)
  2. ที่ปรึกษาคณะทำงาน 4 ท่าน (NEW)
  3. แผน Output/Outcome 1-3-5 ปี (ตัด Soft Power)
  4. อำนาจหน้าที่
  5. งบประมาณ 5 ปี (NEW)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
thin = Side(border_style="thin", color="B0B8C1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Colour palette
C_NAVY   = "1F4E78"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_GREEN  = "E2EFDA"
C_YELLOW = "FFF2CC"
C_ORANGE = "FCE4D6"
C_GRAY   = "F2F7FA"
C_STRIPE = "F8F9FA"
C_RED_SOFT = "FDE8E8"
C_PURPLE = "EDE7F6"
C_TEAL   = "D1FAE5"

def fill(hex_color): return PatternFill("solid", start_color=hex_color)
def font(size=11, bold=False, color="111827", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_cell(ws, row, col, value, bg=C_NAVY, fg="FFFFFF", size=11, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h=h, v="center")
    c.border = BORDER

def data_cell(ws, row, col, value, bg=None, bold=False, h="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color="111827")
    if bg: c.fill = fill(bg)
    c.alignment = align(h=h)
    c.border = BORDER

def merge_title(ws, row, col_start, col_end, value, bg=C_NAVY, fg="FFFFFF", size=14, height=40):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font = Font(name=FONT, size=size, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h="center", v="center")
    c.border = BORDER
    ws.row_dimensions[row].height = height

wb = Workbook()

# ══════════════════════════════════════════════════════
# SHEET 1: คณะทำงาน 17 คน (Final v2 — AICAT ประธาน)
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "คณะทำงาน 17 คน"
ws1.sheet_view.showGridLines = False

merge_title(ws1, 1, 1, 6,
    "คณะทำงานขับเคลื่อนศูนย์ความเป็นเลิศปัญญาประดิษฐ์ด้านอุตสาหกรรมสร้างสรรค์ (COE AI – Creative Economy)")

merge_title(ws1, 2, 1, 6,
    "เจ้าภาพประสาน: สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)  |  แต่งตั้งโดย: คณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติ  |  จัดทำ พ.ค. 2569",
    bg=C_BLUE, size=10, height=22)

headers = ["ลำดับ", "บล็อก", "ตำแหน่งในคณะทำงาน", "ชื่อ–สกุล / ผู้แทน", "สังกัด / หน่วยงาน", "บทบาทและความเชี่ยวชาญ"]
for col, h in enumerate(headers, 1):
    header_cell(ws1, 3, col, h, bg=C_LBLUE, fg="1F2937", size=11)
ws1.row_dimensions[3].height = 32

BLOCK_COLORS = {
    "A": (C_PURPLE, "ผู้นำ"),
    "B": (C_GREEN,  "ภาครัฐ"),
    "C": (C_YELLOW, "ผู้ทรงคุณวุฒิ"),
    "D": (C_ORANGE, "เลขานุการ"),
}

# 17 คน — Final v2
members = [
    # Block A — ผู้นำ (2)
    ("A","ประธานคณะทำงาน",
     "นายสุธัช เจริญผล",
     "นายกสมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "กำหนดทิศทางเชิงยุทธศาสตร์ COE / กำกับภาพรวม\nเจ้าภาพประสานภาครัฐ–เอกชน–วิชาการ"),

    ("A","รองประธานคณะทำงาน",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานส่งเสริมเศรษฐกิจดิจิทัล (depa)",
     "เชื่อมนโยบาย Digital Economy / สนับสนุนโครงสร้างพื้นฐานดิจิทัล\nรองประธานหลัก ทำหน้าที่แทนประธานในที่ประชุม"),

    # Block B — กรรมการรัฐ (6)
    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมวัฒนธรรม กระทรวงวัฒนธรรม",
     "เศรษฐกิจวัฒนธรรม + มรดกทางวัฒนธรรม / นโยบายสื่อสร้างสรรค์\nเชื่อม Creative Economy กับต้นทุนวัฒนธรรมไทย"),

    ("B","กรรมการ",
     "ดร.ชาคริต พิชญางกูร",
     "สำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "Champion นโยบาย AI x Creative Economy (CREATECH)\nMandate จากรองนายกฯ นภินทร (14 พ.ค. 2569)"),

    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมทรัพย์สินทางปัญญา กระทรวงพาณิชย์",
     "AI Copyright / IP Registry / ลิขสิทธิ์เนื้อหา AI-generated\nกำกับกรอบกฎหมาย AI Creative Content"),

    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานนวัตกรรมแห่งชาติ (NIA)",
     "กลไก Grant + Startup Funding สำหรับ Creative AI\nสนับสนุน AI Creative Startup/SME"),

    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สถาบันข้อมูลขนาดใหญ่ (BDI)",
     "Thai Creative AI Dataset / Data Infrastructure\nสร้างฐานข้อมูลภาษาและวัฒนธรรมไทยสำหรับ AI"),

    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมการค้าระหว่างประเทศ (DITP)",
     "Export Creative Content / Bangkok International Content Market\nผลักดันคอนเทนต์ไทยและ AI Tools สู่ตลาด ASEAN"),

    # Block C — ผู้ทรงคุณวุฒิ (7)
    ("C","ผู้ทรงคุณวุฒิ (1)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้กำกับภาพยนตร์ไทย",
     "อุตสาหกรรมภาพยนตร์ + AI Film Production\nผู้กำกับและผู้ผลิตคอนเทนต์ภาพยนตร์ไทย"),

    ("C","ผู้ทรงคุณวุฒิ (2)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้ประกอบการแอนิเมชันและคอมพิวเตอร์กราฟิกไทย (TACGA)",
     "AI ใน Animation / VFX / Virtual Production\nอุตสาหกรรมภาพยนตร์และแอนิเมชันไทย"),

    ("C","ผู้ทรงคุณวุฒิ (3)",
     "(นายก / ผู้แทน)",
     "สมาคมสมาพันธ์โอเพนซอร์สแห่งประเทศไทย",
     "AI Open Source + Foundation Model ไทย\nระบบนิเวศนักพัฒนา Open Source ไทย"),

    ("C","ผู้ทรงคุณวุฒิ (4)",
     "(โปรดระบุชื่อ–นามสกุล)",
     "ผู้เชี่ยวชาญด้านคอมพิวเตอร์กราฟิก (CGI) และวิชวลเอฟเฟกต์",
     "AI ใน CGI / VFX / Production Pipeline\nเทคโนโลยีการผลิตภาพและเสียงระดับ Industry"),

    ("C","ผู้ทรงคุณวุฒิ (5)",
     "(นายก / ผู้แทน)",
     "สมาคมโฆษณาดิจิทัล (DAAT)",
     "Generative AI สำหรับ Digital Advertising / Personalization\nAI ในกระบวนการผลิต Campaign"),

    ("C","ผู้ทรงคุณวุฒิ (6)",
     "(นายก / ผู้แทน)",
     "สมาคม Content Creator (โปรดยืนยันชื่อเต็ม)",
     "Creator Economy / AI Tools สำหรับ Individual Creator\nUGC + Influencer + Short-form Content"),

    ("C","ผู้ทรงคุณวุฒิ (7)",
     "อ.สมชาย (โปรดระบุนามสกุล)",
     "ผู้เชี่ยวชาญด้านกฎหมายทรัพย์สินทางปัญญาและปัญญาประดิษฐ์",
     "AI Copyright Law / IP Framework / จรรยาบรรณ AI Creative\nร่างแนวปฏิบัติและมาตรฐาน AI Creative ฉบับไทย"),

    # Block D — เลขานุการ (2)
    ("D","กรรมการและเลขานุการ",
     "(โปรดระบุชื่อ–ผู้แทนระดับปฏิบัติ สดช.)",
     "สำนักงานคณะกรรมการดิจิทัลเพื่อเศรษฐกิจและสังคม (สดช.)",
     "รายงานผลกลับ คกก.เฉพาะด้านฯ AI แห่งชาติ\nบริหารวาระประชุม / ติดตามผล / เอกสาร"),

    ("D","กรรมการและเลขานุการร่วม",
     "(โปรดระบุชื่อ–ผู้แทน AICAT)",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานภาคเอกชน/สมาคมในคณะ\nบริหารงานเครือข่ายและกิจกรรม COE"),
]

assert len(members) == 17, f"Expected 17 members, got {len(members)}"

start_row = 4
for i, (blk, pos, name, org, role) in enumerate(members, 1):
    r = start_row + i - 1
    bg_color, _ = BLOCK_COLORS[blk]
    stripe = C_STRIPE if i % 2 == 0 else None
    row_bg = bg_color if i <= 2 or i >= 16 else stripe

    data_cell(ws1, r, 1, i, bg=row_bg, bold=True, h="center")
    data_cell(ws1, r, 2, blk, bg=row_bg, bold=True, h="center")
    data_cell(ws1, r, 3, pos, bg=row_bg, bold=(blk in ("A","D")))
    data_cell(ws1, r, 4, name, bg=row_bg)
    data_cell(ws1, r, 5, org, bg=row_bg)
    data_cell(ws1, r, 6, role, bg=row_bg)
    ws1.row_dimensions[r].height = 52

for col, w in zip(range(1, 7), [5, 6, 28, 22, 36, 42]):
    ws1.column_dimensions[get_column_letter(col)].width = w

# Legend
leg_row = start_row + 17 + 1
ws1.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=6)
c = ws1.cell(row=leg_row, column=1,
    value="📌 หมายเหตุ: ชื่อที่ระบุว่า 'โปรดระบุ' = ต้องกรอกก่อนส่ง สดช.  |  "
          "🔴 ดร.ชาคริต ต้อง pre-sync ก่อน (เปลี่ยนจากประธานเดิม → กรรมการ)  |  "
          "🟡 ชื่อ-นามสกุลผู้ทรงคุณวุฒิ Block C หลายรายรอยืนยัน  |  "
          "ดูที่ปรึกษาคณะทำงาน 4 ท่านใน Sheet ถัดไป")
c.font = Font(name=FONT, size=9, italic=True, color="374151")
c.fill = fill("FEF9C3")
c.alignment = align(h="left", v="center", wrap=True)
c.border = BORDER
ws1.row_dimensions[leg_row].height = 40
ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 2: ที่ปรึกษาคณะทำงาน 4 ท่าน
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("ที่ปรึกษาคณะทำงาน")
ws2.sheet_view.showGridLines = False

merge_title(ws2, 1, 1, 5,
    "ที่ปรึกษาคณะทำงาน COE AI Creative Economy (นอก 17 ที่นั่งหลัก)")

merge_title(ws2, 2, 1, 5,
    "สถานะ: ผู้ทรงคุณวุฒิที่เข้าร่วมประชุมตามวาระ — ไม่มีสิทธิ์ออกเสียง แต่ลงชื่อในคำสั่งฯ ได้",
    bg=C_BLUE, size=10, height=22)

adv_headers = ["ลำดับ", "ตำแหน่ง", "ชื่อ / ผู้แทน", "ความเชี่ยวชาญ / สังกัด", "บทบาทใน COE"]
for col, h in enumerate(adv_headers, 1):
    header_cell(ws2, 3, col, h, bg=C_LBLUE, fg="1F2937")
ws2.row_dimensions[3].height = 32

advisors = [
    ("ป.1", "ที่ปรึกษาคณะทำงาน",
     "(ผู้บริหาร / ผู้แทน)",
     "บริษัท เวิร์คพอยท์ เอ็นเทอร์เทนเมนต์ จำกัด (มหาชน)",
     "Mass Content Industry / TV + Digital Distribution\nที่ปรึกษาด้านอุตสาหกรรมคอนเทนต์ระดับ Mainstream"),

    ("ป.2", "ที่ปรึกษาคณะทำงาน",
     "(โปรดระบุชื่อ–นามสกุล)",
     "ผู้ทรงคุณวุฒิด้านเทคโนโลยีการผลิตภาพยนตร์ (Film Production Technology)",
     "AI Production Pipeline / Virtual Production\nที่ปรึกษาด้านเทคโนโลยีการผลิตเชิงลึก"),

    ("ป.3", "ที่ปรึกษาคณะทำงาน",
     "(โปรดระบุชื่อ–นามสกุล + มหาวิทยาลัย)",
     "นักวิจัย / อาจารย์ด้านปัญญาประดิษฐ์",
     "AI Research / Academic Network\nที่ปรึกษาเชิงวิชาการและงานวิจัย"),

    ("ป.4", "ที่ปรึกษาคณะทำงาน",
     "(ผู้บริหาร / ผู้แทน)",
     "บริษัท บิทคับ ออนไลน์ จำกัด (Bitkub Online)",
     "Digital Asset / NFT สำหรับเศรษฐกิจสร้างสรรค์\nที่ปรึกษาด้านสินทรัพย์ดิจิทัลและตลาดทุนใหม่"),
]

for i, (no, pos, name, org, role) in enumerate(advisors, 1):
    r = 3 + i
    bg = C_TEAL if i % 2 == 1 else None
    data_cell(ws2, r, 1, no, bg=bg, bold=True, h="center")
    data_cell(ws2, r, 2, pos, bg=bg, bold=True)
    data_cell(ws2, r, 3, name, bg=bg)
    data_cell(ws2, r, 4, org, bg=bg)
    data_cell(ws2, r, 5, role, bg=bg)
    ws2.row_dimensions[r].height = 55

for col, w in zip(range(1, 6), [7, 24, 28, 38, 42]):
    ws2.column_dimensions[get_column_letter(col)].width = w

# Note
note_row = 3 + len(advisors) + 2
ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
nc = ws2.cell(row=note_row, column=1,
    value="กลไกที่ปรึกษา: ตามแบบคณะทำงานราชการ — เข้าร่วมประชุมได้ตามวาระ ให้คำปรึกษาเชิงเทคนิค "
          "ลงนามเป็นที่ปรึกษาในคำสั่งแต่งตั้ง แต่ไม่นับใน 17 ที่นั่งหลัก และไม่มีสิทธิ์ออกเสียงในมติประชุม")
nc.font = Font(name=FONT, size=9, italic=True, color="374151")
nc.fill = fill("FEF9C3")
nc.alignment = align(h="left", v="center", wrap=True)
nc.border = BORDER
ws2.row_dimensions[note_row].height = 42
ws2.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 3: แผน Output/Outcome 1-3-5 ปี (ตัด Soft Power)
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("แผน Output-Outcome 1-3-5 ปี")
ws3.sheet_view.showGridLines = False

merge_title(ws3, 1, 1, 5,
    "เป้าหมาย Output / Outcome ของ COE AI Creative Economy  |  ระยะ 1 ปี / 3 ปี / 5 ปี")
merge_title(ws3, 2, 1, 5,
    "ปี 1 = Foundation 2569–70  |  ปี 3 = Scale 2571–72  |  ปี 5 = Leadership 2573–74",
    bg=C_BLUE, size=10, height=22)

h3 = ["มิติ (Pillar)", "ปี 1 — Foundation\n(Output)", "ปี 3 — Scale\n(Outcome)", "ปี 5 — Leadership\n(Impact)", "KPI / ตัวชี้วัด"]
for col, h in enumerate(h3, 1):
    header_cell(ws3, 3, col, h, bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[3].height = 42

# === PILLARS — Soft Power REMOVED, replaced with "เศรษฐกิจวัฒนธรรม / อิทธิพลทางวัฒนธรรม" ===
pillars = [
    ("1) Hub &\nโครงสร้างพื้นฐาน",
     "• ตั้ง CoE Hub ร่วมกับ CEA/TCDC\n• เครือข่ายพันธมิตร MOU ≥ 10 องค์กร\n• แพลตฟอร์มกลาง AI Creative",
     "• ขยาย Hub ครอบคลุม 5 ภูมิภาค\n• เครือข่าย ASEAN ≥ 5 ประเทศ",
     "• ไทยเป็น AI Creative Hub ของ ASEAN\n• Hub เชื่อมต่อกับ Global AI Creative Network",
     "จำนวน MOU / จำนวน Hub / Active Users"),

    ("2) Talent &\nการศึกษา",
     "• เปิดหลักสูตร AI Creative Production ≥ 2 หลักสูตร\n• ผู้ผ่านอบรม ≥ 500 คน",
     "• Talent Pool ≥ 1,000 คน (ผ่านรับรองมาตรฐาน)\n• หลักสูตรขยายสู่ ≥ 10 สถาบัน",
     "• Talent Pipeline ป้อนอุตสาหกรรมต่อเนื่อง\n• Thai AI Creative Talent ทำงานในเวทีโลก",
     "จำนวนผู้ผ่านอบรม / อัตราการจ้างงาน"),

    ("3) Sandbox &\nนำร่อง",
     "• Sandbox 3–5 โปรเจกต์\n  (Animation / Music / Content / Ads)\n• สร้าง Use Case + Baseline",
     "• Scale จาก Pilot สู่ Commercial ≥ 10 โปรเจกต์\n• AI Creative Startup ≥ 20 ราย",
     "• AI Creative Startup/SME ที่ scale ในตลาด\n• Unicorn ด้าน Creative AI 1–2 ราย",
     "จำนวน Startup / มูลค่าเชิงพาณิชย์"),

    ("4) Thai Creative\nAI Dataset",
     "• Thai Creative AI Dataset Phase 1\n  (ภาษา วัฒนธรรม ทรัพยากรสร้างสรรค์ไทย)",
     "• Thai Creative AI Dataset Phase 2–3\n  (ขยาย sector + open access ภาคอุตสาหกรรม)",
     "• Dataset เป็น ASEAN Creative AI Benchmark\n• Thai AI Tools/IP ใช้งานใน ≥ 5 ประเทศ",
     "จำนวน Dataset / ประเทศที่ใช้งาน"),

    ("5) มาตรฐาน\n& กฎหมาย IP",
     "• กรอบนโยบาย AI Copyright & IP (เบื้องต้น)\n• แนวปฏิบัติ AI Creative Content ฉบับไทย",
     "• ประกาศใช้มาตรฐานวิชาชีพ AI Creative ระดับชาติ\n• จรรยาบรรณ AI Creative ฉบับไทย (official)",
     "• ไทยเป็นผู้นำมาตรฐาน AI Creative ใน ASEAN\n• ส่งออกโมเดลนโยบายให้ประเทศอื่นอ้างอิง",
     "จำนวนนโยบายที่ประกาศใช้ / การยอมรับสากล"),

    ("6) ต้นทุนผลิต\n& ผลิตภาพ",
     "• Baseline ต้นทุนผลิตคอนเทนต์ปัจจุบัน\n• เครื่องมือ AI ที่ผู้ประกอบการเข้าถึงได้",
     "• ต้นทุนผลิตคอนเทนต์ลดลง ≥ 30%\n  (เทียบ Baseline ปีที่ 1)",
     "• AI Adoption ≥ 80% ของผู้ประกอบการ Creative\n• ผลิตภาพเพิ่มขึ้นอย่างมีนัยสำคัญ",
     "% ต้นทุนลด / % AI Adoption"),

    # ↓↓↓ Pillar 7 — Soft Power → "เวทีระดับโลก / อิทธิพลทางวัฒนธรรม"
    ("7) Export &\nเวทีระดับโลก",
     "• ศึกษาตลาด ASEAN + จัดทำ Export Roadmap\n• เข้าร่วมงาน Bangkok International Content Market",
     "• เริ่ม export AI Creative Content ไทยสู่ ASEAN\n  ผ่าน DITP + Bangkok Content Market",
     "• Thai AI Creative Content ในเวทีระดับโลก\n  (รางวัล/ยอดชมระดับนานาชาติ)\n• อิทธิพลทางวัฒนธรรมไทยขยายตัวในเวทีโลก",
     "มูลค่าส่งออก / รางวัลนานาชาติ"),

    ("8) ผลกระทบ\nเชิงเศรษฐกิจ\n(KPI หลัก)",
     "• Baseline: Creative Economy = 8.78% ของ GDP\n  (CEA รายงาน พ.ค. 2569)",
     "• เพิ่มการจ้างงานสาขาสร้างสรรค์ ≥ 10,000 ตำแหน่ง\n• Creative AI Startup สร้างรายได้วัดได้",
     "• Creative Economy → 12%+ ของ GDP\n  โดย AI เป็นตัวขับเคลื่อนหลัก\n• ผู้ประกอบการใช้ AI ≥ 80% ของ sector",
     "% GDP / % AI Adoption\n(KPI หลัก 2 ตัว)"),
]

start3 = 4
for i, (pillar, y1, y3, y5, kpi) in enumerate(pillars):
    r = start3 + i
    is_last = (i == len(pillars) - 1)
    pillar_bg = "FDE8E8" if is_last else None
    data_cell(ws3, r, 1, pillar, bg=pillar_bg, bold=is_last)
    data_cell(ws3, r, 2, y1, bg=C_GREEN)
    data_cell(ws3, r, 3, y3, bg=C_YELLOW)
    data_cell(ws3, r, 4, y5, bg=C_ORANGE)
    data_cell(ws3, r, 5, kpi, bold=is_last)
    ws3.row_dimensions[r].height = 95 if i < len(pillars)-1 else 115

# Vision footer — Soft Power → "เศรษฐกิจวัฒนธรรมและอุตสาหกรรมสร้างสรรค์"
vr = start3 + len(pillars) + 1
ws3.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=5)
vc = ws3.cell(row=vr, column=1,
    value='วิสัยทัศน์ 5 ปี: ทำให้ประเทศไทยเป็น "ASEAN Hub of AI-powered Creative Economy" '
          '— ผลักดันเศรษฐกิจวัฒนธรรมและอุตสาหกรรมสร้างสรรค์ไทยสู่เวทีโลก '
          'ด้วยกำลังคน เทคโนโลยี IP และระบบนิเวศที่พร้อม')
vc.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
vc.fill = fill(C_NAVY)
vc.alignment = align(h="center", v="center")
vc.border = BORDER
ws3.row_dimensions[vr].height = 50

for col, w in zip(range(1, 6), [18, 34, 34, 34, 26]):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = "B4"

# ══════════════════════════════════════════════════════
# SHEET 4: อำนาจหน้าที่
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet("อำนาจหน้าที่")
ws4.sheet_view.showGridLines = False

merge_title(ws4, 1, 1, 3,
    "(ร่าง) อำนาจหน้าที่คณะทำงานขับเคลื่อน COE AI Creative Economy")
merge_title(ws4, 2, 1, 3,
    "อ้างอิง: ข้อ 2.5 คำสั่ง คกก.ดิจิทัลเพื่อเศรษฐกิจและสังคมแห่งชาติ ที่ 1/2569",
    bg=C_BLUE, size=10, height=22)

header_cell(ws4, 3, 1, "ข้อ", bg=C_LBLUE, fg="1F2937")
header_cell(ws4, 3, 2, "อำนาจหน้าที่", bg=C_LBLUE, fg="1F2937")
header_cell(ws4, 3, 3, "หน่วยงานรับผิดชอบหลัก", bg=C_LBLUE, fg="1F2937")
ws4.row_dimensions[3].height = 30

duties = [
    ("1",
     "ศึกษา วิเคราะห์ และจัดทำแผนการขับเคลื่อนศูนย์ความเป็นเลิศปัญญาประดิษฐ์ด้านอุตสาหกรรมสร้างสรรค์ (COE AI Creative Economy) ให้สอดคล้องกับแผนปฏิบัติการด้านปัญญาประดิษฐ์แห่งชาติ และยุทธศาสตร์ชาติ 20 ปี",
     "AICAT + CEA + สดช."),
    ("2",
     "ส่งเสริม สนับสนุน และขับเคลื่อนการนำปัญญาประดิษฐ์มาประยุกต์ใช้ในอุตสาหกรรมสร้างสรรค์ของประเทศไทย ครอบคลุมสาขา ภาพยนตร์ แอนิเมชัน ดนตรี เกม โฆษณา แฟชั่น ออกแบบ คอนเทนต์ดิจิทัล และผู้สร้างสรรค์เนื้อหา",
     "TACGA, DAAT, สมาคมผู้กำกับฯ, Content Creator, AICAT"),
    ("3",
     "ประสานความร่วมมือระหว่างภาครัฐ ภาคเอกชน สถาบันการศึกษา และองค์กรระหว่างประเทศ เพื่อสร้างระบบนิเวศ (Ecosystem) ที่เอื้อต่อการพัฒนา AI Creative Economy ของไทย",
     "AICAT, CEA, depa, DITP"),
    ("4",
     "กำกับดูแลและผลักดันการจัดทำมาตรฐานวิชาชีพ จรรยาบรรณ และกรอบกฎหมายด้านปัญญาประดิษฐ์สำหรับอุตสาหกรรมสร้างสรรค์ไทย รวมถึงแนวปฏิบัติด้านลิขสิทธิ์ผลงาน AI-generated",
     "อ.สมชาย, กรมทรัพย์สินทางปัญญา, กรมส่งเสริมวัฒนธรรม"),
    ("5",
     "ส่งเสริมการพัฒนากำลังคนและ Talent Pipeline ด้าน AI Creative ผ่านหลักสูตรการศึกษา การฝึกอบรม และ Sandbox สำหรับผู้ประกอบการ โดยมีเป้าหมาย Talent Pool ≥ 1,000 คน ภายในปีที่ 3",
     "CEA, NIA, AICAT, ที่ปรึกษาด้านวิชาการ"),
    ("6",
     "สนับสนุนการสร้าง Thai Creative AI Dataset และโครงสร้างพื้นฐานข้อมูลเพื่อรองรับการพัฒนา AI สำหรับอุตสาหกรรมสร้างสรรค์ไทย",
     "BDI, depa, สมาคมโอเพนซอร์ส"),
    ("7",
     "ส่งเสริมการส่งออกผลงานและเครื่องมือ AI Creative ของไทยสู่ตลาด ASEAN และระดับโลก ผ่านงาน Bangkok International Content Market และความร่วมมือกับ DITP",
     "DITP, DAAT, AICAT"),
    ("8",
     "ติดตาม ประเมิน และรายงานผลการดำเนินงานต่อคณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติ เป็นระยะตามที่กำหนด (ทุก 6 เดือน)",
     "สดช. (เลขานุการ) + AICAT (เลขานุการร่วม)"),
    ("9",
     "แต่งตั้งคณะอนุทำงาน คณะทำงานเฉพาะกิจ หรือเชิญผู้ทรงคุณวุฒิเพิ่มเติมเข้าร่วมประชุมเพื่อดำเนินการในเรื่องใดเรื่องหนึ่งตามที่คณะทำงานมอบหมาย",
     "ประธาน + คณะทำงาน"),
    ("10",
     "ปฏิบัติหน้าที่อื่นตามที่คณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติมอบหมาย",
     "—"),
]

start4 = 4
for i, (num, duty, resp) in enumerate(duties):
    r = start4 + i
    bg = C_STRIPE if i % 2 == 0 else None
    data_cell(ws4, r, 1, num, bg=bg, bold=True, h="center")
    data_cell(ws4, r, 2, duty, bg=bg)
    data_cell(ws4, r, 3, resp, bg=bg)
    ws4.row_dimensions[r].height = 66

for col, w in zip(range(1, 4), [5, 72, 28]):
    ws4.column_dimensions[get_column_letter(col)].width = w
ws4.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 5: งบประมาณ 5 ปี (NEW)
# ══════════════════════════════════════════════════════
ws5 = wb.create_sheet("งบประมาณ 5 ปี")
ws5.sheet_view.showGridLines = False

merge_title(ws5, 1, 1, 7,
    "ประมาณการงบประมาณ COE AI Creative Economy  |  5 ปี (2569–2574)  |  หน่วย: ล้านบาท")
merge_title(ws5, 2, 1, 7,
    "Scenario แนะนำ: Mid (744 MB / 5 ปี)  |  Bottom-up จากกิจกรรมจริง × อัตราอ้างอิงตลาดไทย",
    bg=C_BLUE, size=10, height=22)

budget_headers = ["มิติ (Pillar)", "ปี 1", "ปี 2", "ปี 3", "ปี 4", "ปี 5", "รวม 5 ปี"]
for col, h in enumerate(budget_headers, 1):
    header_cell(ws5, 3, col, h, bg=C_LBLUE, fg="1F2937")
ws5.row_dimensions[3].height = 30

budget_rows = [
    ("Hub & Operations (CEA/TCDC co-host)", 12, 12, 15, 15, 18, 72),
    ("หลักสูตร + Training", 8, 12, 20, 20, 20, 80),
    ("Sandbox Projects (3–5 โปรเจกต์/ปี)", 20, 25, 30, 25, 20, 120),
    ("Thai Creative AI Dataset (Phase 1→3)", 20, 25, 30, 20, 15, 110),
    ("นโยบาย / กฎหมาย / มาตรฐาน", 5, 8, 10, 8, 5, 36),
    ("Startup Grants (NIA)", 0, 10, 30, 40, 30, 110),
    ("Export + ASEAN (DITP)", 3, 8, 15, 20, 25, 71),
    ("Events / Showcase / PR", 8, 10, 15, 15, 20, 68),
    ("Administration / เลขานุการ", 5, 5, 5, 5, 5, 25),
    ("ASEAN / Global Partnerships", 2, 5, 10, 15, 20, 52),
]

start5 = 4
for i, row_data in enumerate(budget_rows):
    r = start5 + i
    bg = C_STRIPE if i % 2 == 0 else None
    data_cell(ws5, r, 1, row_data[0], bg=bg, bold=True)
    for col_idx, val in enumerate(row_data[1:], 2):
        data_cell(ws5, r, col_idx, val, bg=bg, h="right")
    ws5.row_dimensions[r].height = 28

# Total row
total_r = start5 + len(budget_rows)
totals = [sum(row[i] for row in budget_rows) for i in range(1, 7)]
data_cell(ws5, total_r, 1, "รวมต่อปี", bg=C_NAVY, bold=True)
ws5.cell(row=total_r, column=1).font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
for col_idx, val in enumerate(totals, 2):
    data_cell(ws5, total_r, col_idx, val, bg=C_NAVY, h="right", bold=True)
    ws5.cell(row=total_r, column=col_idx).font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
ws5.row_dimensions[total_r].height = 32

# Scenario block
sr = total_r + 2
ws5.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=7)
sc = ws5.cell(row=sr, column=1,
    value="📊 3 Scenarios: Conservative 400 MB  |  Mid (แนะนำ) 744 MB  |  Ambitious 1,100 MB")
sc.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
sc.fill = fill(C_BLUE)
sc.alignment = align(h="center", v="center")
sc.border = BORDER
ws5.row_dimensions[sr].height = 30

# Funding sources
fr = sr + 2
header_cell(ws5, fr, 1, "แหล่งงบประมาณ", bg=C_LBLUE, fg="1F2937")
header_cell(ws5, fr, 2, "สัดส่วน", bg=C_LBLUE, fg="1F2937")
header_cell(ws5, fr, 3, "ล้านบาท", bg=C_LBLUE, fg="1F2937")
ws5.merge_cells(start_row=fr, start_column=4, end_row=fr, end_column=7)
header_cell(ws5, fr, 4, "กลไก", bg=C_LBLUE, fg="1F2937")
ws5.row_dimensions[fr].height = 30

funding = [
    ("งบแผ่นดิน (CEA, สดช., วธ.)", "40%", 298, "งบดำเนินงานปกติ + โครงการพิเศษ"),
    ("depa Digital Economy Fund", "20%", 149, "Grant สำหรับ Digital/AI projects"),
    ("NIA Innovation Fund", "15%", 112, "Startup grant + R&D grant"),
    ("ภาคเอกชน / อุตสาหกรรม", "10%", 74, "Co-investment, in-kind, sponsorship"),
    ("ทุนต่างประเทศ / MOU", "10%", 74, "Research grants, bilateral programs"),
    ("DITP / BOI incentives", "5%", 37, "Export support, investment promotion"),
]

for i, (src, pct, amt, mech) in enumerate(funding, 1):
    r = fr + i
    bg = C_STRIPE if i % 2 == 1 else None
    data_cell(ws5, r, 1, src, bg=bg, bold=True)
    data_cell(ws5, r, 2, pct, bg=bg, h="center")
    data_cell(ws5, r, 3, amt, bg=bg, h="right")
    ws5.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    data_cell(ws5, r, 4, mech, bg=bg)
    ws5.row_dimensions[r].height = 26

# ROI note
roi_r = fr + len(funding) + 2
ws5.merge_cells(start_row=roi_r, start_column=1, end_row=roi_r, end_column=7)
rc = ws5.cell(row=roi_r, column=1,
    value="💰 ROI คาดการณ์: ทุก 1 บาทที่ลงทุน → GDP เพิ่มขึ้น ~820 บาท  "
          "(GDP Creative Economy 8.78% → 12% = +610,000 MB / ต้นทุนลด 30% = +500,000 MB)")
rc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
rc.fill = fill("059669")
rc.alignment = align(h="center", v="center", wrap=True)
rc.border = BORDER
ws5.row_dimensions[roi_r].height = 40

for col, w in zip(range(1, 8), [38, 10, 10, 10, 10, 10, 14]):
    ws5.column_dimensions[get_column_letter(col)].width = w
ws5.freeze_panes = "B4"

# ══════════════════════════════════════════════════════
out = "COE_AI_Creative_Economy_v3.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
print(f"   Sheets: {wb.sheetnames}")

# Verify no Soft Power
import re
soft_power_count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if re.search(r'soft\s*power', cell.value, re.IGNORECASE):
                    soft_power_count += 1
                    print(f"   ⚠ Found Soft Power in {sheet_name}: {cell.value[:60]}")

if soft_power_count == 0:
    print("   ✅ ไม่พบคำว่า 'Soft Power' ในไฟล์ — ผ่าน")
else:
    print(f"   ❌ พบคำว่า 'Soft Power' {soft_power_count} ตำแหน่ง")
