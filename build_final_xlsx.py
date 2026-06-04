"""สร้าง Excel ฉบับสมบูรณ์ — COE AI Creative Economy
3 Sheets:
  1. คณะทำงาน 17 คน (Final)
  2. แผน Output/Outcomes 1-3-5 ปี (Final)
  3. อำนาจหน้าที่ (Draft)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
thin = Side(border_style="thin", color="B0B8C1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Colour palette
C_NAVY   = "1F4E78"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_GREEN  = "E2EFDA"
C_YELLOW = "FFF2CC"
C_ORANGE = "FCE4D6"
C_GRAY   = "F2F7FA"
C_STRIPE = "F8F9FA"
C_RED_SOFT = "FDE8E8"
C_PURPLE = "EDE7F6"

def fill(hex_color): return PatternFill("solid", start_color=hex_color)
def font(size=11, bold=False, color="111827", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_cell(ws, row, col, value, bg=C_NAVY, fg="FFFFFF", size=11, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h=h, v="center")
    c.border = BORDER

def data_cell(ws, row, col, value, bg=None, bold=False, h="left", size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, color="111827")
    if bg: c.fill = fill(bg)
    c.alignment = align(h=h)
    c.border = BORDER

def merge_title(ws, row, col_start, col_end, value, bg=C_NAVY, fg="FFFFFF", size=14, height=40):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font = Font(name=FONT, size=size, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h="center", v="center")
    c.border = BORDER
    ws.row_dimensions[row].height = height

wb = Workbook()

# ══════════════════════════════════════════════════════
# SHEET 1: คณะทำงาน 17 คน
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "คณะทำงาน 17 คน"
ws1.sheet_view.showGridLines = False

merge_title(ws1, 1, 1, 6,
    "คณะทำงานขับเคลื่อนศูนย์ความเป็นเลิศปัญญาประดิษฐ์ด้านอุตสาหกรรมสร้างสรรค์ (COE AI – Creative Economy)")

merge_title(ws1, 2, 1, 6,
    "เจ้าภาพประสาน: สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)  |  แต่งตั้งโดย: คณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติ  |  จัดทำ พ.ค. 2569",
    bg=C_BLUE, size=10, height=22)

headers = ["ลำดับ", "บล็อก", "ตำแหน่งในคณะทำงาน", "ชื่อ–สกุล / ผู้แทน", "สังกัด / หน่วยงาน", "บทบาทและความเชี่ยวชาญ"]
for col, h in enumerate(headers, 1):
    header_cell(ws1, 3, col, h, bg=C_LBLUE, fg="1F2937", size=11)
ws1.row_dimensions[3].height = 32

# Block colours
BLOCK_COLORS = {
    "A": (C_PURPLE, "ผู้นำ"),
    "B": (C_GREEN,  "ภาครัฐ"),
    "C": (C_YELLOW, "ผู้ทรงคุณวุฒิ"),
    "D": (C_ORANGE, "เลขานุการ"),
}

members = [
    # (block, ตำแหน่ง, ชื่อ, สังกัด, บทบาท)
    ("A","ประธานคณะทำงาน",
     "ดร.ชาคริต พิชญางกูร",
     "ผู้อำนวยการสำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "กำหนดทิศทางเชิงยุทธศาสตร์ของ COE / กำกับภาพรวม\nChampion นโยบาย AI x Creative Economy (CREATECH)"),

    ("A","รองประธาน คนที่ 1",
     "(โปรดระบุชื่อ–นายก AICAT)",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานภาคเอกชน/สมาคม / ขับเคลื่อนงานปฏิบัติ\nผู้ประสานงานหลักระหว่างภาครัฐ–เอกชน"),

    ("A","รองประธาน คนที่ 2",
     "(โปรดระบุชื่อ–ผอ.depa หรือผู้แทน)",
     "สำนักงานส่งเสริมเศรษฐกิจดิจิทัล (depa)",
     "เชื่อมนโยบาย Digital Economy / สนับสนุนโครงสร้างพื้นฐานดิจิทัล"),

    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมวัฒนธรรม กระทรวงวัฒนธรรม",
     "Soft Power + Cultural Heritage / นโยบายสื่อสร้างสรรค์\nเชื่อม Creative Economy กับต้นทุนวัฒนธรรมไทย"),

    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมทรัพย์สินทางปัญญา กระทรวงพาณิชย์",
     "AI Copyright / IP Registry / ลิขสิทธิ์เนื้อหา AI-generated\nกำกับกรอบกฎหมาย AI Creative Content"),

    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สำนักงานนวัตกรรมแห่งชาติ (NIA)",
     "กลไก Grant + Startup Funding สำหรับ Creative AI\nสนับสนุน AI Creative Startup/SME"),

    ("B","กรรมการ",
     "(ผู้อำนวยการ / ผู้แทน)",
     "สถาบันข้อมูลขนาดใหญ่ (BDI)",
     "Thai Creative AI Dataset / Data Infrastructure\nสร้างฐานข้อมูลภาษาและวัฒนธรรมไทยสำหรับ AI"),

    ("B","กรรมการ",
     "(อธิบดี / ผู้แทน)",
     "กรมส่งเสริมการค้าระหว่างประเทศ (DITP)",
     "Export Creative Content / Bangkok International Content Market\nผลักดันคอนเทนต์ไทยและ AI Tools สู่ตลาด ASEAN"),

    ("C","ผู้ทรงคุณวุฒิ (1)",
     "(นายก / ผู้แทน)",
     "สมาคมโฆษณาดิจิทัล (DAAT)",
     "Generative AI สำหรับ Digital Advertising / Personalization\nAI ในกระบวนการผลิต Campaign"),

    ("C","ผู้ทรงคุณวุฒิ (2)",
     "(นายก / ผู้แทน)",
     "สมาคมผู้ประกอบการแอนิเมชันและคอมพิวเตอร์กราฟิกไทย (TACGA)",
     "AI ใน Animation / VFX / Virtual Production\nอุตสาหกรรมภาพยนตร์และแอนิเมชันไทย"),

    ("C","ผู้ทรงคุณวุฒิ (3)",
     "(ผู้แทน)",
     "สภาอุตสาหกรรมแห่งประเทศไทย (FTI) — กลุ่มดิจิทัล/Creative",
     "ตัวแทนภาคอุตสาหกรรมสร้างสรรค์รวม\nเชื่อม SME Creative กับนโยบายภาคอุตสาหกรรม"),

    ("C","ผู้ทรงคุณวุฒิ (4)",
     "อ.สมชาย (โปรดระบุนามสกุล)",
     "ผู้เชี่ยวชาญด้านกฎหมายทรัพย์สินทางปัญญาและปัญญาประดิษฐ์",
     "AI Copyright Law / IP Framework / จรรยาบรรณ AI Creative\nร่างแนวปฏิบัติและมาตรฐาน AI Creative ฉบับไทย"),

    ("C","ผู้ทรงคุณวุฒิ (5)",
     "(ผู้อำนวยการ / ผู้แทน)",
     "ศูนย์ส่งเสริมอุตสาหกรรมสร้างสรรค์และดนตรี (MACI)\nคณะดุริยางคศาสตร์ มหาวิทยาลัยศิลปากร",
     "Music Industry + AI Music Production\nTalent Pipeline ด้าน Creative Arts / Academic Research"),

    ("C","ผู้ทรงคุณวุฒิ (6)",
     "(นายก / ผู้แทน)",
     "สมาคม Content Creator (โปรดระบุชื่อสมาคม)",
     "Creator Economy / AI Tools สำหรับ Individual Creator\nUGC + Influencer + Short-form Content"),

    ("D","กรรมการและเลขานุการ",
     "(โปรดระบุชื่อ–ผู้แทนระดับปฏิบัติ CEA)",
     "สำนักงานส่งเสริมเศรษฐกิจสร้างสรรค์ (CEA)",
     "บริหารวาระการประชุม / จัดทำรายงาน / ติดตามผล\nประสานงานหน่วยงานภาครัฐในคณะ"),

    ("D","กรรมการและเลขานุการร่วม",
     "(โปรดระบุชื่อ–ผู้แทนระดับปฏิบัติ สดช.)",
     "สำนักงานคณะกรรมการดิจิทัลเพื่อเศรษฐกิจและสังคม (สดช.)",
     "รายงานผลกลับ คกก.เฉพาะด้านฯ AI แห่งชาติ\nประสานงานเชิงนโยบาย AI ระดับชาติ"),

    ("D","กรรมการและผู้ช่วยเลขานุการ",
     "(โปรดระบุชื่อ–ผู้แทน AICAT)",
     "สมาคมสร้างสรรค์ปัญญาประดิษฐ์ไทย (AICAT)",
     "ประสานภาคเอกชน/สมาคมในคณะ\nบริหารงานเครือข่ายและกิจกรรม COE"),
]

assert len(members) == 17

start_row = 4
for i, (blk, pos, name, org, role) in enumerate(members, 1):
    r = start_row + i - 1
    bg_color, _ = BLOCK_COLORS[blk]
    stripe = C_STRIPE if i % 2 == 0 else None
    row_bg = bg_color if i <= 3 or i >= 15 else stripe

    data_cell(ws1, r, 1, i, bg=row_bg, bold=True, h="center")
    data_cell(ws1, r, 2, blk, bg=row_bg, bold=True, h="center")
    data_cell(ws1, r, 3, pos, bg=row_bg, bold=(blk in ("A","D")))
    data_cell(ws1, r, 4, name, bg=row_bg)
    data_cell(ws1, r, 5, org, bg=row_bg)
    data_cell(ws1, r, 6, role, bg=row_bg)
    ws1.row_dimensions[r].height = 52

# Block divider labels (merge col B for each block visually via color already done)
for col, w in zip(range(1, 7), [5, 6, 28, 22, 36, 42]):
    ws1.column_dimensions[get_column_letter(col)].width = w

# Legend row
leg_row = start_row + 17 + 1
ws1.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=6)
c = ws1.cell(row=leg_row, column=1,
    value="📌 หมายเหตุ: ชื่อที่ระบุว่า 'โปรดระบุ' = ต้องกรอกก่อนส่ง สดช.  |  "
          "🟡 ประธาน + รองประธาน 2 รอยืนยัน pre-sync  |  "
          "ผู้ทรงคุณวุฒิ (4) รอนามสกุล อ.สมชาย  |  ผู้ทรงคุณวุฒิ (6) รอชื่อเต็มสมาคม Content Creator")
c.font = Font(name=FONT, size=9, italic=True, color="374151")
c.fill = fill("FEF9C3")
c.alignment = align(h="left", v="center", wrap=True)
c.border = BORDER
ws1.row_dimensions[leg_row].height = 28
ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
# SHEET 2: แผน Output/Outcomes 1-3-5 ปี
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("แผน Output-Outcomes 1-3-5 ปี")
ws2.sheet_view.showGridLines = False

merge_title(ws2, 1, 1, 5,
    "เป้าหมาย Output / Outcomes ของ COE AI Creative Economy  |  ระยะ 1 ปี / 3 ปี / 5 ปี")
merge_title(ws2, 2, 1, 5,
    "ปี 1 = Foundation 2569–70  |  ปี 3 = Scale 2571–72  |  ปี 5 = Leadership 2573–74",
    bg=C_BLUE, size=10, height=22)

h2 = ["มิติ (Pillar)", "ปี 1 — Foundation\n(Output)", "ปี 3 — Scale\n(Outcome)", "ปี 5 — Leadership\n(Impact)", "KPI / ตัวชี้วัด"]
for col, h in enumerate(h2, 1):
    header_cell(ws2, 3, col, h, bg=C_LBLUE, fg="1F2937")
ws2.row_dimensions[3].height = 42

pillars = [
    ("1) Hub &\nโครงสร้างพื้นฐาน",
     "• ตั้ง CoE Hub ร่วมกับ CEA/TCDC\n• เครือข่ายพันธมิตร MOU ≥ 10 องค์กร\n• แพลตฟอร์มกลาง AI Creative",
     "• ขยาย Hub ครอบคลุม 5 ภูมิภาค\n• เครือข่าย ASEAN ≥ 5 ประเทศ",
     "• ไทยเป็น AI Creative Hub ของ ASEAN\n• Hub เชื่อมต่อกับ Global AI Creative Network",
     "จำนวน MOU / จำนวน Hub / Active Users"),

    ("2) Talent &\nการศึกษา",
     "• เปิดหลักสูตร AI Creative Production ≥ 2 หลักสูตร\n• ผู้ผ่านอบรม ≥ 500 คน",
     "• Talent Pool ≥ 1,000 คน (ผ่านรับรองมาตรฐาน)\n• หลักสูตรขยายสู่ ≥ 10 สถาบัน",
     "• Talent Pipeline ป้อนอุตสาหกรรมต่อเนื่อง\n• Thai AI Creative Talent ทำงาน World Stage",
     "จำนวนผู้ผ่านอบรม / อัตราการจ้างงาน"),

    ("3) Sandbox &\nนำร่อง",
     "• Sandbox 3–5 โปรเจกต์\n  (Animation / Music / Content / Ads)\n• สร้าง Use Case + Baseline",
     "• Scale จาก Pilot สู่ Commercial ≥ 10 โปรเจกต์\n• AI Creative Startup ≥ 20 ราย",
     "• AI Creative Startup/SME ที่ scale ในตลาด\n• Unicorn ด้าน Creative AI 1–2 ราย",
     "จำนวน Startup / มูลค่าเชิงพาณิชย์"),

    ("4) Thai Creative\nAI Dataset",
     "• Thai Creative AI Dataset Phase 1\n  (ภาษา วัฒนธรรม ทรัพยากรสร้างสรรค์ไทย)",
     "• Thai Creative AI Dataset Phase 2–3\n  (ขยาย sector + open access ภาคอุตสาหกรรม)",
     "• Dataset เป็น ASEAN Creative AI Benchmark\n• Thai AI Tools/IP ใช้งานใน ≥ 5 ประเทศ",
     "จำนวน Dataset / ประเทศที่ใช้งาน"),

    ("5) มาตรฐาน\n& กฎหมาย IP",
     "• กรอบนโยบาย AI Copyright & IP (เบื้องต้น)\n• แนวปฏิบัติ AI Creative Content ฉบับไทย",
     "• ประกาศใช้มาตรฐานวิชาชีพ AI Creative ระดับชาติ\n• จรรยาบรรณ AI Creative ฉบับไทย (official)",
     "• ไทยเป็นผู้นำมาตรฐาน AI Creative ใน ASEAN\n• ส่งออกโมเดลนโยบายให้ประเทศอื่นอ้างอิง",
     "จำนวนนโยบายที่ประกาศใช้ / การยอมรับสากล"),

    ("6) ต้นทุนผลิต\n& ผลิตภาพ",
     "• Baseline ต้นทุนผลิตคอนเทนต์ปัจจุบัน\n• เครื่องมือ AI ที่ผู้ประกอบการเข้าถึงได้",
     "• ต้นทุนผลิตคอนเทนต์ลดลง ≥ 30%\n  (เทียบ Baseline ปีที่ 1)",
     "• AI Adoption ≥ 80% ของผู้ประกอบการ Creative\n• ผลิตภาพเพิ่มขึ้นอย่างมีนัยสำคัญ",
     "% ต้นทุนลด / % AI Adoption"),

    ("7) Export &\nSoft Power",
     "• ศึกษาตลาด ASEAN + จัดทำ Export Roadmap\n• เข้าร่วมงาน Bangkok International Content Market",
     "• เริ่ม export AI Creative Content ไทยสู่ ASEAN\n  ผ่าน DITP + Bangkok Content Market",
     "• Thai AI Creative Content ใน World Stage\n  (รางวัล/ยอดชมระดับนานาชาติ)\n• Soft Power Index ไทยเพิ่มขึ้น",
     "มูลค่าส่งออก / รางวัลนานาชาติ"),

    ("8) ผลกระทบ\nเชิงเศรษฐกิจ\n(KPI หลัก)",
     "• Baseline: Creative Economy = 8.78% ของ GDP\n  (CEA รายงาน พ.ค. 2569)",
     "• เพิ่มการจ้างงานสาขาสร้างสรรค์ ≥ 10,000 ตำแหน่ง\n• Creative AI Startup สร้างรายได้วัดได้",
     "• Creative Economy → 12%+ ของ GDP\n  โดย AI เป็นตัวขับเคลื่อนหลัก\n• ผู้ประกอบการใช้ AI ≥ 80% ของ sector",
     "% GDP / % AI Adoption\n(KPI หลัก 2 ตัว)"),
]

start2 = 4
fills2 = [C_GREEN, C_YELLOW, C_ORANGE]  # ปี1, ปี3, ปี5
for i, (pillar, y1, y3, y5, kpi) in enumerate(pillars):
    r = start2 + i
    is_last = (i == len(pillars) - 1)
    pillar_bg = "FDE8E8" if is_last else None  # highlight KPI หลัก
    data_cell(ws2, r, 1, pillar, bg=pillar_bg, bold=is_last)
    data_cell(ws2, r, 2, y1, bg=C_GREEN)
    data_cell(ws2, r, 3, y3, bg=C_YELLOW)
    data_cell(ws2, r, 4, y5, bg=C_ORANGE)
    data_cell(ws2, r, 5, kpi, bold=is_last)
    ws2.row_dimensions[r].height = 90 if i < len(pillars)-1 else 110

# Vision footer
vr = start2 + len(pillars) + 1
ws2.merge_cells(start_row=vr, start_column=1, end_row=vr, end_column=5)
vc = ws2.cell(row=vr, column=1,
    value='วิสัยทัศน์ 5 ปี: ทำให้ประเทศไทยเป็น "ASEAN Hub of AI-powered Creative Economy" '
          '— ผลักดัน Soft Power ไทยสู่เวทีโลก ด้วยกำลังคน เทคโนโลยี IP และระบบนิเวศที่พร้อม')
vc.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
vc.fill = fill(C_NAVY)
vc.alignment = align(h="center", v="center")
vc.border = BORDER
ws2.row_dimensions[vr].height = 44

for col, w in zip(range(1, 6), [18, 34, 34, 34, 26]):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = "B4"

# ══════════════════════════════════════════════════════
# SHEET 3: อำนาจหน้าที่
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("อำนาจหน้าที่")
ws3.sheet_view.showGridLines = False

merge_title(ws3, 1, 1, 3,
    "(ร่าง) อำนาจหน้าที่คณะทำงานขับเคลื่อน COE AI Creative Economy")
merge_title(ws3, 2, 1, 3,
    "อ้างอิง: ข้อ 2.5 คำสั่ง คกก.ดิจิทัลเพื่อเศรษฐกิจและสังคมแห่งชาติ ที่ 1/2569",
    bg=C_BLUE, size=10, height=22)

header_cell(ws3, 3, 1, "ข้อ", bg=C_LBLUE, fg="1F2937")
header_cell(ws3, 3, 2, "อำนาจหน้าที่", bg=C_LBLUE, fg="1F2937")
header_cell(ws3, 3, 3, "หน่วยงานรับผิดชอบหลัก", bg=C_LBLUE, fg="1F2937")
ws3.row_dimensions[3].height = 30

duties = [
    ("1",
     "ศึกษา วิเคราะห์ และจัดทำแผนการขับเคลื่อนศูนย์ความเป็นเลิศปัญญาประดิษฐ์ด้านอุตสาหกรรมสร้างสรรค์ (COE AI Creative Economy) ให้สอดคล้องกับแผนปฏิบัติการด้านปัญญาประดิษฐ์แห่งชาติ และยุทธศาสตร์ชาติ 20 ปี",
     "CEA + สดช. + AICAT"),
    ("2",
     "ส่งเสริม สนับสนุน และขับเคลื่อนการนำปัญญาประดิษฐ์มาประยุกต์ใช้ในอุตสาหกรรมสร้างสรรค์ของประเทศไทย ครอบคลุมสาขา ภาพยนตร์ แอนิเมชัน ดนตรี เกม โฆษณา แฟชั่น ออกแบบ และคอนเทนต์ดิจิทัล",
     "TACGA, DAAT, MACI, AICAT"),
    ("3",
     "ประสานความร่วมมือระหว่างภาครัฐ ภาคเอกชน สถาบันการศึกษา และองค์กรระหว่างประเทศ เพื่อสร้างระบบนิเวศ (Ecosystem) ที่เอื้อต่อการพัฒนา AI Creative Economy ของไทย",
     "AICAT, CEA, depa, DITP"),
    ("4",
     "กำกับดูแลและผลักดันการจัดทำมาตรฐานวิชาชีพ จรรยาบรรณ และกรอบกฎหมายด้านปัญญาประดิษฐ์สำหรับอุตสาหกรรมสร้างสรรค์ไทย รวมถึงแนวปฏิบัติด้านลิขสิทธิ์ผลงาน AI-generated",
     "อ.สมชาย, กรมทรัพย์สินทางปัญญา, FTI"),
    ("5",
     "ส่งเสริมการพัฒนากำลังคนและ Talent Pipeline ด้าน AI Creative ผ่านหลักสูตรการศึกษา การฝึกอบรม และ Sandbox สำหรับผู้ประกอบการ โดยมีเป้าหมาย Talent Pool ≥ 1,000 คน ภายในปีที่ 3",
     "MACI ม.ศิลปากร, NIA, CEA"),
    ("6",
     "สนับสนุนการสร้าง Thai Creative AI Dataset และโครงสร้างพื้นฐานข้อมูลเพื่อรองรับการพัฒนา AI สำหรับอุตสาหกรรมสร้างสรรค์ไทย",
     "BDI, NECTEC, CEA"),
    ("7",
     "ติดตาม ประเมิน และรายงานผลการดำเนินงานต่อคณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติ เป็นระยะตามที่กำหนด",
     "สดช. (เลขานุการร่วม)"),
    ("8",
     "แต่งตั้งคณะอนุกรรมการ คณะทำงานเฉพาะกิจ หรือคณะที่ปรึกษาเพื่อดำเนินการในเรื่องใดเรื่องหนึ่งตามที่คณะทำงานมอบหมาย",
     "ประธาน + คณะทำงาน"),
    ("9",
     "ปฏิบัติหน้าที่อื่นตามที่คณะกรรมการเฉพาะด้านการขับเคลื่อนแผนด้านปัญญาประดิษฐ์แห่งชาติมอบหมาย",
     "—"),
]

start3 = 4
for i, (num, duty, resp) in enumerate(duties):
    r = start3 + i
    bg = C_STRIPE if i % 2 == 0 else None
    data_cell(ws3, r, 1, num, bg=bg, bold=True, h="center")
    data_cell(ws3, r, 2, duty, bg=bg)
    data_cell(ws3, r, 3, resp, bg=bg)
    ws3.row_dimensions[r].height = 66

for col, w in zip(range(1, 4), [5, 72, 28]):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = "A4"

# ══════════════════════════════════════════════════════
out = "COE_AI_Creative_Economy_v2.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
